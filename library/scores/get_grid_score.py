import os
import sys

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)
print(PROJECT_PATH)

import numpy as np
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils.cell import get_column_letter
from opexebo.analysis import grid_score
from library.maps.get_autocorrelation import get_autocorrelation
from library.maps.get_rate_map import get_rate_map
from library.scores.shuffle_spikes import shuffle_spikes

def compute_grid_score(occupancy_map: np.ndarray, ts: np.ndarray, pos_x: np.ndarray,
                       pos_y: np.ndarray, pos_t: np.ndarray, arena_size: tuple,
                       spikex: np.ndarray, spikey: np.ndarray, kernlen: int, std: int):

    '''
        Computes the grid score of neuron given spike data.

        Params:
            occupancy_map (np.ndarray):
                A 2D numpy array enconding subjects position over entire experiment.
            ts (np.ndarray):
                Spike time stamp array
            pos_x, pos_y, pos_t (np.ndarray):
                Arrays of x,y  coordinate positions as well as timestamps of movement respectively
            arena_size (tuple):
                Dimensions of arena
            spikex, spikey (np.ndarray):
                x and y coordinates of spike events respectively
            kenrnlen, std (int):
                kernel size and standard deviation of kernel for convolutional smoothing.
    '''

    # Extract and flatten spike arrays
    unshuffled_spike_xy = np.zeros((2,len(ts)))
    unshuffled_spike_xy[0] = spikex.flatten()
    unshuffled_spike_xy[1] = spikey.flatten()

    # Compute ratemaps, autocorrelation, and finally grid_score
    rate_map_smooth, rate_map_raw = get_rate_map(pos_x, pos_y, pos_t, arena_size, spikex, spikey, kernlen, std)
    autocorr_map = get_autocorrelation(rate_map_smooth,pos_x,pos_y,arena_size)
    grid_score_object = grid_score(autocorr_map)
    true_grid_score = grid_score_object[0]

    return true_grid_score

def grid_score_shuffle(self, occupancy_map: np.ndarray, arena_size: tuple, ts: np.ndarray,
                       pos_x: np.ndarray, pos_y: np.ndarray, pos_t: np.ndarray, kernlen: int, std: int, **kwargs) -> list:

    '''
        Shuffles position and spike data prior to computing grid scores. Shuffling
        allows us to determine if the probability of a given score is random, or
        demonstrates a meaningful association in the data.

        Params:
            occupancy_map (np.ndarray):
                A 2D numpy array enconding subjects position over entire experiment.
            arena_size (tuple):
                Dimensions of arena
            ts (np.ndarray):
                Array of spike event times
            pos_x, pos_y, pos_t (np.ndarray):
                Arrays of x, y  coordinate positions as well as timestamps of movement respectively
            spikex, spikey (np.ndarray):
                x and y coordinates of spike events respectively
            kernlen, std (int):
                kernel size and standard deviation for convolutional smoothing

        **kwargs:
            xsheet: xlwings excel sheet

        Returns:
            list: grid_scores
            --------
            grid_scores: List of 100 grid scores (1 score per shuffle)
    '''

    grid_scores = []
    shuffled_spike_xy = np.zeros((2,len(ts)))

    # If an excel sheet is passed, set reference
    s = kwargs.get('xsheet',None)
    row_index =  kwargs.get('row_index',None)
    cell_number = kwargs.get('cell_number',None)
    column_index = 2

    # Shuffle spike data
    shuffled_spikes = shuffle_spikes(self, ts, pos_x, pos_y, pos_t)

    # For each set of shuffled data, compute grid score
    for element in shuffled_spikes:
        rate_map_smooth, _ = get_rate_map(pos_x, pos_y, pos_t, arena_size, element[0], element[1], kernlen, std)
        autocorr_map = get_autocorrelation(rate_map_smooth,pos_x,pos_y,arena_size)
        grid_score_object = grid_score(autocorr_map)

        # Populate the excel sheet with the scores
        if s != None and row_index != None and cell_number != None:
            if grid_score_object != None:
                grid_scores.append(float(grid_score_object[0]))
                if s[get_column_letter((row_index-1)*5 + 1) + str(1)] != 'C_' + str(cell_number) +'_BorderShuffle_Top':
                    s[get_column_letter(row_index) + str(1)] = 'C_' + str(cell_number) +'_GridShuffle'
                    s[get_column_letter(row_index) + str(column_index)] = grid_score_object[0]
                    column_index += 1
                    ColumnDimension(s, bestFit=True)
                else:
                    s[get_column_letter(row_index*5) + str(1)] = 'C_' + str(cell_number) +'_GridShuffle'
                    s[get_column_letter(row_index*5) + str(column_index)] = grid_score_object[0]
                    column_index += 1
                    ColumnDimension(s, bestFit=True)

        else:
            if grid_score_object != None:
                grid_scores.append(float(grid_score_object[0]))

    return grid_scores
