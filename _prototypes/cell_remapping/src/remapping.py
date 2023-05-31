import os, sys
import numpy as np
from scipy.spatial.distance import cdist
from pyemd import emd
import itertools
from openpyxl import load_workbook
import re
import pandas as pd
import matplotlib.pyplot as plt
import copy
from scipy import stats, ndimage
from skimage.measure import block_reduce
import cv2
import xlsxwriter

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from library.hafting_spatial_maps import SpatialSpikeTrain2D
from _prototypes.cell_remapping.src.rate_map_plots import plot_obj_remapping, plot_regular_remapping, plot_fields_remapping, plot_shuffled_regular_remapping, plot_matched_sesssion_waveforms
from _prototypes.cell_remapping.src.wasserstein_distance import sliced_wasserstein, single_point_wasserstein, pot_sliced_wasserstein, compute_centroid_remapping, _get_ratemap_bucket_midpoints
from _prototypes.cell_remapping.src.masks import make_object_ratemap, check_disk_arena, flat_disk_mask, generate_grid, _sample_grid
from library.maps import map_blobs
from scripts.batch_map.batch_map import batch_map 
from _prototypes.cell_remapping.src.settings import obj_output, centroid_output, tasks, session_comp_categories, regular_output, context_output, variations, temporal_output
from scripts.batch_map.LEC_naming import LEC_naming_format, extract_name_lec
from _prototypes.cell_remapping.src.MEC_naming import MEC_naming_format, extract_name_mec
from _prototypes.cell_remapping.src.LC_naming import LC_naming_format, extract_name_lc
from library.shuffle_spikes import shuffle_spikes

def compute_modified_zscore(x, ref_dist):
    # Compute the median of the data
    median = np.median(ref_dist)

    # Compute the Median Absolute Deviation (MAD)
    mad = stats.median_abs_deviation(ref_dist, scale='normal')

    # Compute the Modified Z-score, adjusting for division by zero
    modified_zscore = 0.6745 * (x - median) / (mad if mad else 1)

    return modified_zscore, median, mad   
                    
def _check_single_format(filename, format, fxn):
    print(filename, format, fxn)
    if re.match(str(format), str(filename)) is not None:
        return fxn(filename)
    
def _single_shuffled_sample(norm, raw, settings):
    # norm, raw = spatial_spike_train.get_map('rate').get_rate_map(new_size = settings['ratemap_dims'][0], shuffle=True)
    if settings['normalizeRate']:
        rate_map = norm
        # rate_map = rate_map / np.sum(rate_map)
    else:
        rate_map = raw   

    # check no nans
    assert np.isnan(rate_map).any() == False, "rate map contains nans pre downsampling"

    if settings['downsample']:
        rate_map = _downsample(rate_map, settings['downsample_factor'])

    # check no nans
    assert np.isnan(rate_map).any() == False, "rate map contains nans post downsampling"
    

    return rate_map 

def _downsample(img, downsample_factor):
    downsampled = block_reduce(img, downsample_factor) 
    return downsampled
    

def compute_remapping(study, settings, data_dir):

    # c = 0
    isStart = True
    context_paths = {}
    # batch_map(study, tasks, ratemap_size=settings['ratemap_dims'][0])

    # if settings['hasObject'] or settings['runFields']:    
    # max_centroid_count, blobs_dict, shuffled_ratemap_dict, shuffled_sample_dict = _aggregate_cell_info(study, settings)

    ratemap_size = settings['ratemap_dims'][0]

        # non_disk_ids = list(itertools.product(np.arange(0, ratemap_size), np.arange(0, ratemap_size)))
        # fake_map = np.random.random((ratemap_size, ratemap_size))
        # if settings['downsample']:
        #     fake_map = _downsample(fake_map, settings['downsample_factor'])
        #     # fake_map = np.random.random((ratemap_size // settings['downsample_factor'], ratemap_size // settings['downsample_factor']))
        # fake_map = flat_disk_mask(fake_map)
        # row, col = np.where(~np.isnan(fake_map))
        # valid_disk_indices = np.array([row, col]).T

        # non_disk_grid_sample = _sample_grid(non_disk_ids, settings['grid_sample_threshold'])
        # disk_grid_sample = _sample_grid(valid_disk_indices, settings['grid_sample_threshold'])?

        # non_disk_grid_sample = generate_grid()

        

    # centroid_dict = copy.deepcopy(centroid_output)
    # regular_dict = copy.deepcopy(regular_output)
    # context_dict = copy.deepcopy(context_output)
    # obj_dict = copy.deepcopy(obj_output)

    for animal in study.animals:
        
        max_centroid_count, blobs_dict, shuffled_ratemap_dict, shuffled_sample_dict = _aggregate_cell_info(animal, settings)

        # if settings['useMatchedCut']:
        #     # get largest possible cell id
        #     max_matched_cell_count = len(animal.sessions[sorted(list(animal.sessions.keys()))[-1]].get_cell_data()['cell_ensemble'].cells)
        # else:
        #     max_matched_cell_count = max(list(map(lambda x: max(animal.sessions[x].get_cell_data()['cell_ensemble'].get_label_ids()), animal.sessions)))

        max_matched_cell_count = max(list(map(lambda x: max(animal.sessions[x].get_cell_data()['cell_ensemble'].get_label_ids()), animal.sessions)))

        # len(session) - 1 bcs thats number of comparisons. e.g. 3 session: ses1-ses2, ses2-ses3 so 2 distances will be given for remapping
        remapping_distances = np.zeros((len(list(animal.sessions.keys()))-1, max_matched_cell_count))
        # remapping_indices = [[] for k in range(max_matched_cell_count)]
        # remapping_session_ids = [[] for k in range(max_matched_cell_count)]

        # for every existing cell id across all sessions
        for k in range(int(max_matched_cell_count)):
            centroid_dict = copy.deepcopy(centroid_output)
            regular_dict = copy.deepcopy(regular_output)
            context_dict = copy.deepcopy(context_output)
            obj_dict = copy.deepcopy(obj_output)
            temporal_dict = copy.deepcopy(temporal_output)

            cell_label = k + 1
            print('Cell ' + str(cell_label))

            # prev ratemap
            prev = None
            curr_spikes = None 
            curr_spike_times = None
            curr_shuffled_temporal = None
            prev_spike_times = None
            prev_id = None
            curr_shuffled = None
            cell_session_appearances = []
            
            # for every session
            for i in range(len(list(animal.sessions.keys()))):
                seskey = 'session_' + str(i+1)
                print(seskey)
                ses = animal.sessions[seskey]
                path = ses.session_metadata.file_paths['tet']
                fname = path.split('/')[-1].split('.')[0]

                if settings['disk_arena']: 
                    cylinder = True
                else:
                    cylinder, _ = check_disk_arena(fname)
                    if not cylinder:
                        print('Not cylinder for {}'.format(fname))
                

                if settings['naming_type'] == 'LEC':
                    group, name = extract_name_lec(fname)
                    formats = LEC_naming_format[group][name][settings['type']]
                elif settings['naming_type'] == 'MEC':
                    name = extract_name_mec(fname)
                    formats = MEC_naming_format
                elif settings['naming_type'] == 'LC':
                    name = extract_name_lc(fname)
                    formats = LC_naming_format

                for format in list(formats.keys()):
                    checked = _check_single_format(fname, format, formats[format])
                    if checked is not None:
                        break
                    else:
                        continue
                
                stim, depth, name, date = checked

                ### TEMPORARY WAY TO READ OBJ LOC FROM FILE NAME ###
                if settings['hasObject']:
                    # object_location = _read_location_from_file(path, cylinder, true_var)

                    object_location = stim
                    
                    if object_location != 'NO' and '_' not in object_location:
                        object_location = int(object_location)

                ensemble = ses.get_cell_data()['cell_ensemble']

                # Check if cell id we're iterating through is present in the ensemble of this sessions
                if cell_label in ensemble.get_cell_label_dict():
                    cell = ensemble.get_cell_by_id(cell_label)
                    cell_session_appearances.append(cell)

                    # spatial_spike_train = ses.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj})

                    # spatial_spike_train = cell.stats_dict['cell_stats']['spatial_spike_train']
                    spatial_spike_train = cell.stats_dict['spatial_spike_train'] 

                    rate_map_obj = spatial_spike_train.get_map('rate')

                    if settings['normalizeRate']:
                        rate_map, _ = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])
                    else:
                        _, rate_map = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])

                    assert rate_map.shape == (settings['ratemap_dims'][0], settings['ratemap_dims'][1]), 'Wrong ratemap shape {} vs settings shape {}'.format(rate_map.shape, (settings['ratemap_dims'][0], settings['ratemap_dims'][1]))
                    
                    # Disk mask ratemap
                    if cylinder:
                        curr = flat_disk_mask(rate_map)
                        if settings['downsample']:
                            curr_ratemap = _downsample(rate_map, settings['downsample_factor'])
                            curr_ratemap = flat_disk_mask(curr_ratemap)
                        else:
                            curr_ratemap = curr
                        row, col = np.where(~np.isnan(curr_ratemap))
                        disk_ids = np.array([row, col]).T
                    else:
                        curr = rate_map
                        if settings['downsample']:
                            curr_ratemap = _downsample(rate_map, settings['downsample_factor']) 
                        else:
                            curr_ratemap = curr
                        disk_ids = None
                    
                    curr_cell = cell
                    curr_spatial_spike_train = spatial_spike_train
                    curr_key = seskey
                    curr_path = ses.session_metadata.file_paths['tet'].split('/')[-1].split('.')[0]
                    curr_id = str(animal.animal_id) + '_' + str(seskey) + '_' + str(cell.cluster.cluster_label)

                    y, x = curr.shape
                    h, w = rate_map_obj.arena_size
                    print('SEARCHTHIS')
                    print(h, w)
                    bin_area = h/y * w/x

                    # If object used in experiment 
                    if settings['hasObject']:

                        _, _, labels, centroids, field_sizes = blobs_dict[curr_id]

                        # print(np.unique(labels), centroids, field_sizes)
                        assert len(np.unique(labels)[1:]) == len(centroids) == len(field_sizes), 'Mismatch in number of labels, centroids and field sizes'
                                    
                        labels_copy = np.copy(labels)
                        labels_copy[np.isnan(curr)] = 0
                        c_count = len(np.unique(labels_copy)) - 1

                        # make all field labels = 1
                        labels_copy[labels_copy != 0] = 1
                        val_r, val_c = np.where(labels_copy == 1)
                        # cumulative_coverage = np.max(field_sizes)
                        cumulative_coverage = len(np.where(labels_copy != 0)[0])/len(np.where(~np.isnan(curr))[0])
                        cumulative_area = len(np.where(labels_copy == 1)[0])
                        cumulative_rate = np.sum(curr[val_r, val_c])

                        height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(rate_map_obj.arena_size, y, x)

                        # from collections import Counter
                        # print(Counter(labels.flatten()))
                        # if settings['downsample']:
                            # labels = _downsample(labels, settings['downsample_factor'])
                            # dfactor = settings['downsample_factor']
                            # labels_resized = cv2.resize(labels, (labels.shape[1]//dfactor, labels.shape[0]//dfactor), interpolation=cv2.INTER_NEAREST)
                            # curr_labels = np.array(labels_resized).astype('uint8')
                            # print(Counter(curr_labels.flatten()))
                            # check no nans
                            # assert np.isnan(curr_labels).all() == False, 'Downsampling error curr labels {}'.format(np.unique(curr_labels))
                            # curr_labels = flat_disk_mask(curr_labels)
                            # print(Counter(curr_labels.flatten()))
                            # re-add nans for disk mask
                            # rnan, cnan = np.where(np.isnan(labels))
                            # curr_labels = np.copy(labels_resized).astype('float')
                            # curr_labels[rnan, cnan] = np.nan

                            # assert np.unique(curr_labels[~np.isnan(curr_labels)]).all() == np.unique(labels).all(), 'Downsampling error curr labels {} vs labels {}'.format(np.unique(curr_labels), np.unique(labels))
                        # else:
                            # curr_labels = labels
                        curr_labels = labels

                        # n_repeats = 100

                        # if cylinder:
                        #     # row, col = np.where(~np.isnan(curr))
                        #     # disk_ids = np.array([row, col]).T
                        #     # # resampled_positions = list(map(lambda x: disk_ids[np.random.choice(np.arange(len(disk_ids)),size=1)[0]] , np.arange(0, n_repeats)))
                        #     # # resampled_positions = disk_ids[np.random.choice(np.arange(len(disk_ids)),size=n_repeats, replace=True)]
                        #     # resampled_positions = _sample_grid(disk_ids, 3)
                        #     resampled_positions = disk_grid_sample
                        # else:
                        #     # options = list(itertools.product(np.arange(0, curr.shape[0]), np.arange(0, curr.shape[1])))
                        #     # # resampled_positions = list(map(lambda x: options[np.random.choice(np.arange(len(options)),size=1)[0]] , np.arange(0, n_repeats)))
                        #     # # resampled_positions = np.array(options)[np.random.choice(np.arange(len(options)),size=n_repeats, replace=True)]
                        #     # resampled_positions = _sample_grid(options, 3)
                        #     resampled_positions = non_disk_grid_sample

                        resampled_positions = generate_grid(rate_map_obj.arena_size[0], rate_map_obj.arena_size[1], 
                                                                settings['spacing'], is_hexagonal=settings['hexagonal'], is_cylinder=cylinder)

                        print('Resampled positions: ', len(resampled_positions), len(resampled_positions[0]))

                        obj_map_dict = {}
                        for var in variations:

                            if settings['downsample']:
                                object_ratemap, object_pos = make_object_ratemap(var, new_size=int(settings['ratemap_dims'][0]/settings['downsample_factor']))
                            else:
                                object_ratemap, object_pos = make_object_ratemap(var, new_size=settings['ratemap_dims'][0])

                            if cylinder:
                                object_ratemap = flat_disk_mask(object_ratemap)
                                # disk_ids = valid_disk_indices
                                # ids where not nan
                            #     row, col = np.where(~np.isnan(object_ratemap))
                            #     disk_ids = np.array([row, col]).T
                            # else:
                                # disk_ids = None

                            obj_map_dict[var] = [object_ratemap, object_pos, disk_ids]
                        
                        # ['whole', 'field', 'bin', 'centroid']
                        for obj_score in settings['object_scores']:
                            unq_labels = np.unique(labels)[1:]
                            for lid in range(len(unq_labels)):
                                label_id = unq_labels[lid]
                                true_object_pos = None
                                true_object_ratemap = None
                                # # Possible object locations (change to get from settings)
                                # variations = [0,90,180,270,'no']
                                # compute object remapping for every object position, actual object location is store alongside wass for each object ratemap
                                resampled_wass = None
                                for var in variations:
                                    obj_wass_key = 'obj_wass_' + str(var)
                                    obj_vector_key = 'obj_vec_' + str(var)
                                    obj_quantile_key = 'obj_q_' + str(var)

                                    object_ratemap, object_pos, disk_ids = obj_map_dict[var]

                                    if var == object_location:
                                        true_object_pos = object_pos
                                        true_object_ratemap = object_ratemap

                                    if isinstance(object_pos, dict):
                                        obj_x = width_bucket_midpoints[object_pos['x']]
                                        obj_y = height_bucket_midpoints[object_pos['y']]
                                    else:
                                        obj_y = height_bucket_midpoints[object_pos[0]]
                                        obj_x = width_bucket_midpoints[object_pos[1]] 

                                    y, x = curr.shape
                                    height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(curr_spatial_spike_train.arena_size, y, x)

                                    # EMD on norm/unnorm ratemap + object map for OBJECT remapping
                                    if obj_score == 'whole' and lid == 0:
                                        # if var == 'no':
                                        #     obj_wass = pot_sliced_wasserstein(coord_buckets, coord_buckets, source_weights, target_weights, n_projections=settings['n_projections'])
                                        # else:

                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, ids=disk_ids)

                                        # if obj_wass != obj_wass:
                                        #     print('NAN OBJ WASS')
                                        #     print(curr_ratemap.shape, disk_ids)
                                        #     print(curr_ratemap)
                                        # n_repeats = 1000
                                        # if disk_ids is not None:
                                        #     resampled_positions = list(map(lambda x: disk_ids[np.random.choice(np.arange(len(disk_ids)),size=1)[0]] , np.arange(0, n_repeats)))
                                        # else:
                                        #     options = list(itertools.product(np.arange(0, y), np.arange(0, x)))
                                        #     resampled_positions = list(map(lambda x: options[np.random.choice(np.arange(len(options)),size=1)[0]] , np.arange(0, n_repeats)))
                                        
                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, ids=disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        # if quantile != quantile:
                                        #     print('NAN QUANTILE')
                                        #     print(resampled_wass)
                                        #     print(obj_wass)

                                        # find row col of peak firing rate bin
                                        r, c = np.where(curr == np.nanmax(curr))
                                        # print(obj_y, obj_x, r, c)
                                        r = height_bucket_midpoints[r[0]]
                                        c = width_bucket_midpoints[c[0]]
                                        mag = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((r, c)))
                                        pt1 = (obj_y, obj_x)
                                        pt2 = (r, c)
                                        angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

                                    elif obj_score == 'field':

                                        # field ids is ids of binary field/map blolb

                                        # TAKE ONLY MAIN FIELD --> already sorted by size
                                        row, col = np.where(curr_labels == label_id)
                                        # print('LABEL ID: ', label_id)
                                        # print('ROW COL: ', row, col)
                                        field_ids = np.array([row, col]).T
                                        if cylinder:
                                            # take ids that are both in disk and in field
                                            print('IT IS A CYLINDER, TAKING ONLY IDS IN FIELD AND IN DISK')
                                            field_disk_ids = np.array([x for x in field_ids if x in disk_ids])
                                        else:
                                            field_disk_ids = field_ids
                                        
                                        # print('FIELD DISK IDS: ', field_disk_ids)
                                        # print('DISK IDS: ', disk_ids)
                                        # print('FIELD IDS: ', field_ids)
                                        # if var == 'no':
                                        #     obj_wass = pot_sliced_wasserstein(coord_buckets, coord_buckets[field_disk_ids], source_weights, target_weights[field_disk_ids], n_projections=settings['n_projections'])
                                        # else:
                                        # NEED TO NORMALIZE FIELD NOT WHOLE MAP
                                        # save total mass in firing field, save area for each field
                                        # remember long format
                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, ids=field_disk_ids)

                                        # if label_id == 1 or label_id == 2:
                                        #     print('LABEL ID: ', label_id)
                                        #     print(len(row), len(col), len(field_disk_ids), len(disk_ids))
                                        #     if len(row) == 0:
                                        #         print('NO FIELD FOR LABEL: ', label_id)
                                        #     print(np.unique(curr_labels), np.unique(labels))
                                        #     print(curr_labels)
                                        #     print(obj_wass)
                                        #     print(np.sum(curr_ratemap), np.sum(curr))
                                        #     print(np.sum(curr_ratemap[disk_ids[:,0],disk_ids[:,1]]), np.sum(curr[disk_ids[:,0],disk_ids[:,1]]))
                                        #     print(np.sum(curr_ratemap[field_disk_ids[:,0], field_disk_ids[:,1]]), np.sum(curr[field_disk_ids[:,0], field_disk_ids[:,1]]))
                                        #     print(curr_ratemap[field_disk_ids[:,0], field_disk_ids[:,1]])
                                    
                                        # n_repeats = 1000
                                        # resampled_positions = list(map(lambda x: field_disk_ids[np.random.choice(np.arange(len(field_disk_ids)),size=1)[0]] , np.arange(0, n_repeats)))
                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, ids=field_disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()


                                        # x,y bins of peak firing rate
                                        # idx = np.where(source_weights == np.max(source_weights))[0]
                                        # r, c = field_disk_ids[idx]
                                        r, c = centroids[0]
                                        # print(r, c)
                                        r = height_bucket_midpoints[int(np.round(r))]
                                        c = width_bucket_midpoints[int(np.round(c))]
                                        mag = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((r, c)))
                                        pt1 = (obj_y, obj_x)
                                        pt2 = (r, c)
                                        angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

                                    elif obj_score == 'spike_density' and lid == 0:
                                        # obj_bin_wass = single_point_wasserstein(object_pos, labels_curr, rate_map_obj.arena_size, ids=field_ids)

                                        # curr_spike_pos_x, curr_spike_pos_y, _ = curr_spatial_spike_train.get_spike_positions()
                                        curr_spike_pos_x, curr_spike_pos_y = curr_spatial_spike_train.spike_x, curr_spatial_spike_train.spike_y
                                        curr_spike_pos_x *= -1
                                        # spiekx an spikey are negative and positive, make positive
                                        curr_spike_pos_x += np.abs(np.min(curr_spike_pos_x))
                                        curr_spike_pos_y += np.abs(np.min(curr_spike_pos_y))
                                
                                        curr_pts = np.array([curr_spike_pos_y, curr_spike_pos_x]).T
                                        # source_pts = np.array([[obj_y], [obj_x]]).T

                                        # obj_wass = pot_sliced_wasserstein(source_pts, curr_pts, n_projections=settings['n_projections'])
                                        obj_wass = single_point_wasserstein(object_pos, curr_ratemap, rate_map_obj.arena_size, density=True, density_map=curr_pts, use_pos_directly=False)

                                        # if var == 'NO':
                                        #     plt.scatter(curr_pts[:,1], curr_pts[:,0], s=2, c='k', alpha=1)
                                        #     plt.show()
                                            
                                        # if disk_ids is not None:
                                        #     resampled_positions = list(map(lambda x: disk_ids[np.random.choice(np.arange(len(disk_ids)),size=1)[0]] , np.arange(0, n_repeats)))
                                        # else:
                                        #     options = list(itertools.product(np.arange(0, y), np.arange(0, x)))
                                        #     resampled_positions = list(map(lambda x: options[np.random.choice(np.arange(len(options)),size=1)[0]] , np.arange(0, n_repeats)))

                                        # n_repeats = 1000
                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_ratemap, rate_map_obj.arena_size, density=True, density_map=curr_pts, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        # avg across all points
                                        r = np.mean(curr_spike_pos_y)
                                        c = np.mean(curr_spike_pos_x)
                                        mag = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((r, c)))
                                        pt1 = (obj_y, obj_x)
                                        pt2 = (r, c)
                                        angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

                                    elif obj_score == 'binary':

                                        # if var == 'no':
                                        #     row, col = np.where(~np.isnan(object_ratemap))
                                        # else:
                                        #     # TAKE ONLY MAIN FIELD --> already sorted by size
                                        row, col = np.where(curr_labels == label_id)
                                        
                                        field_ids = np.array([row, col]).T

                                        if cylinder:
                                            # take ids that are both in disk and in field
                                            print('IT IS A CYLINDER, TAKING ONLY IDS IN FIELD AND IN DISK')
                                            field_disk_ids = np.array([x for x in field_ids if x in disk_ids])
                                        else:
                                            field_disk_ids = field_ids

                                        curr_masked = np.zeros((curr_labels.shape))
                                        curr_masked[field_disk_ids[:,0], field_disk_ids[:,1]] = 1

                                        obj_wass = single_point_wasserstein(object_pos, curr_masked, rate_map_obj.arena_size, ids=field_disk_ids)

                                        # n_repeats = 1000
                                        # resampled_positions = list(map(lambda x: field_disk_ids[np.random.choice(np.arange(len(field_disk_ids)),size=1)[0]] , np.arange(0, n_repeats)))
                                        # compute EMD on resamples
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: single_point_wasserstein(x, curr_masked, rate_map_obj.arena_size, ids=field_disk_ids, use_pos_directly=True), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        r = np.mean(field_disk_ids[:,0])
                                        c = np.mean(field_disk_ids[:,1])
                                        r = height_bucket_midpoints[int(np.round(r))]
                                        c = width_bucket_midpoints[int(np.round(c))]
                                        mag = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((r, c)))
                                        pt1 = (obj_y, obj_x)
                                        pt2 = (r, c)
                                        angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

                                    elif obj_score == 'centroid':

                                        # TAKE ONLY MAIN FIELD --> already sorted by size
                                        main_centroid = centroids[lid]
                                        # row, col = np.where(labels_curr == 1)
                                        # field_ids = np.array([row, col]).T

                                        # euclidean distance between point
                                        obj_wass = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((main_centroid[0],main_centroid[1])))
                                        
                                        if resampled_wass is None:
                                            resampled_wass = list(map(lambda x: np.linalg.norm(np.array((obj_y, obj_x)) - np.array((x[0],x[1]))), resampled_positions))
                                        quantile = (resampled_wass < obj_wass).mean()

                                        # its saving only last angle (NO), either do each or one for true obj pos or one for true and trace
                                        r, c = main_centroid
                                        r = height_bucket_midpoints[int(np.round(r))]
                                        c = width_bucket_midpoints[int(np.round(c))]
                                        mag = np.linalg.norm(np.array((obj_y, obj_x)) - np.array((r, c)))
                                        pt1 = (obj_y, obj_x)
                                        pt2 = (r, c)
                                        angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

                                    if lid != 0 and (obj_score == 'whole' or obj_score == 'spike_density') == True:
                                        pass 
                                    else:
                                        # obj_dict[obj_wass_key].append([obj_wass, obj_field_wass, obj_bin_wass, c_wass])
                                        obj_dict[obj_wass_key].append(obj_wass)
                                        obj_dict[obj_quantile_key].append(quantile)
                                        obj_dict[obj_vector_key].append([pt1, pt2, mag, angle])

                                # if first centroid label, we can save whole map annd spike density scores, if second or later label, we don't want to resave
                                # the whole map and spike density scores
                                if lid != 0 and (obj_score == 'whole' or obj_score == 'spike_density') == True:
                                    pass 
                                else:
                                    val_r, val_c = np.where(labels == label_id)
                                    field_coverage = len(val_r)/len(np.where(~np.isnan(curr))[0])
                                    field_area = len(val_r)
                                    field_rate = np.sum(curr[val_r, val_c])
                                    total_rate = np.sum(curr)
                                    field_peak_rate = np.max(curr[val_r, val_c])

                                    if obj_score == 'whole' or obj_score == 'spike_density':
                                        obj_dict['field_id'].append('all')
                                    else:
                                        obj_dict['field_id'].append(label_id)

                                    obj_dict['score'].append(obj_score)
                                    obj_dict['field_peak_rate'].append(field_peak_rate)
                                    obj_dict['total_rate'].append(total_rate)
                                    obj_dict['field_coverage'].append(field_coverage)
                                    obj_dict['field_area'].append(field_area)
                                    obj_dict['field_rate'].append(field_rate)
                                    obj_dict['cumulative_coverage'].append(cumulative_coverage)
                                    obj_dict['cumulative_area'].append(cumulative_area)
                                    obj_dict['cumulative_rate'].append(cumulative_rate)
                                    obj_dict['field_count'].append(c_count)
                                    obj_dict['bin_area'].append(bin_area[0])
                                    obj_dict['object_location'].append(object_location)
                                    obj_dict['obj_pos'].append((true_object_pos['x'], true_object_pos['y']))
                                    obj_dict['signature'].append(curr_path)
                                    obj_dict['name'].append(name)
                                    obj_dict['date'].append(date)
                                    obj_dict['depth'].append(depth)
                                    obj_dict['unit_id'].append(cell_label)
                                    obj_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    obj_dict['session_id'].append(seskey)
                                    obj_dict['arena_size'].append(curr_spatial_spike_train.arena_size)
                                    obj_dict['cylinder'].append(cylinder)
                                    obj_dict['ratemap_dims'].append(curr.shape)
                                    # obj_dict['grid_sample_threshold'].append(settings['grid_sample_threshold'])
                                    obj_dict['spacing'].append(settings['spacing'])
                                    obj_dict['hexagonal'].append(settings['hexagonal'])
                                    obj_dict['sample_size'].append(len(resampled_positions))

                                    if settings['downsample']:
                                        obj_dict['downsample_factor'].append(settings['downsample_factor'])
                                    else:
                                        obj_dict['downsample_factor'].append(1)

                        if settings['plotObject']:
                            plot_obj_remapping(true_object_ratemap, curr, labels, centroids, obj_dict, data_dir)

                    # If prev ratemap is not None (= we are at session2 or later, session1 has no prev session to compare)
                    if prev is not None and settings['runRegular']:


                        # get x and y pts for spikes in pair of sessions (prev and curr) for a given comparison

                        # prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial_spike_train.get_spike_positions()
                        prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial_spike_train.spike_x, prev_spatial_spike_train.spike_y, prev_spatial_spike_train.new_spike_times
                        prev_pts = np.array([prev_spike_pos_x, prev_spike_pos_y]).T

                        # curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial_spike_train.get_spike_positions()
                        curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial_spike_train.spike_x, curr_spatial_spike_train.spike_y, curr_spatial_spike_train.new_spike_times
                        curr_pts = np.array([curr_spike_pos_x, curr_spike_pos_y]).T

                        if settings['rotate_evening']:
                            if 'evening' in curr_path.lower() or 'rotated' in curr_path.lower():
                                curr_ratemap = ndimage.rotate(curr_ratemap, settings['rotate_angle'])

                                if settings['rotate_angle'] == 90:
                                    curr_pts = np.array([curr_pts[:,1], -curr_pts[:,0]]).T
                                    print('rotating 90 degrees for {}'.format(curr_path))
                                else:
                                    raise ValueError('Rotation angle not supported {}'.format(settings['rotate_angle']))
                            else:
                                print('not rotating for {}'.format(curr_path))

                        y, x = prev_ratemap.shape
                        # find indices of not nan 
                        row_prev, col_prev = np.where(~np.isnan(prev_ratemap))
                        row_curr, col_curr = np.where(~np.isnan(curr_ratemap))

                        print('setting up shuffled samples')
                        # for first map
                        if prev_shuffled is None:  
                            prev_shuffled = shuffled_ratemap_dict[prev_id]
                            prev_shuffled_sample = shuffled_sample_dict[prev_id]
                            # prev_shuffled_samples = list(map(lambda x: _single_shuffled_sample(prev_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                            # if cylinder:
                            #     prev_shuffled_samples = list(map(lambda x: flat_disk_mask(x), prev_shuffled_samples))
                            # prev_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_prev, col_prev))), prev_shuffled_samples))

                        # curr_shuffled_samples = list(map(lambda x: _single_shuffled_sample(curr_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                        # if cylinder:
                        #     curr_shuffled_samples = list(map(lambda x: flat_disk_mask(x), curr_shuffled_samples))   
                        # curr_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_curr, col_curr))), curr_shuffled_samples))
                        curr_shuffled = shuffled_ratemap_dict[curr_id]
                        curr_shuffled_sample = shuffled_sample_dict[curr_id]

                        # assert row_prev.all() == row_curr.all() and col_prev.all() == col_curr.all(), 'Nans in different places'

                        height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(prev_spatial_spike_train.arena_size, y, x)
                        height_bucket_midpoints = height_bucket_midpoints[row_curr]
                        width_bucket_midpoints = width_bucket_midpoints[col_curr]
                        source_weights = np.array(list(map(lambda x, y: prev_ratemap[x,y], row_prev, col_prev)))
                        target_weights = np.array(list(map(lambda x, y: curr_ratemap[x,y], row_curr, col_curr)))
                        source_weights = source_weights / np.sum(source_weights)
                        target_weights = target_weights / np.sum(target_weights)
                        coord_buckets_curr = np.array(list(map(lambda x, y: [height_bucket_midpoints[x],width_bucket_midpoints[y]], row_curr, col_curr)))
                        coord_buckets_prev = np.array(list(map(lambda x, y: [height_bucket_midpoints[x],width_bucket_midpoints[y]], row_prev, col_prev)))

                        print('doing spike density wasserstein')
                        spike_dens_wass = pot_sliced_wasserstein(prev_pts, curr_pts, n_projections=settings['n_projections'])
                            # elif rate_score == 'whole':
                                # This is EMD on whole map for normalized/unnormalized rate remapping
                        print('doing whole map wasserstein')
                        wass = pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, source_weights, target_weights, n_projections=settings['n_projections'])
                        print('doing ref wasserstein')
                        ref_wass_dist = list(map(lambda x, y: pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, x/np.sum(x), y/np.sum(y), n_projections=settings['n_shuffle_projections']), prev_shuffled, curr_shuffled))
                        ref_wass_mean = np.mean(ref_wass_dist)
                        ref_wass_std = np.std(ref_wass_dist)
                        z_score = (wass - ref_wass_mean) / (ref_wass_std)
                        print('doing modified z score')
                        mod_z_score, median, mad = compute_modified_zscore(wass, ref_wass_dist)

                        assert len(ref_wass_dist) == settings['n_repeats'], 'n_repeats does not match length of ref_wass_dist'
   
                        pvalue = stats.t.cdf(z_score, len(ref_wass_dist)-1)
                        mod_pvalue = stats.t.cdf(mod_z_score, len(ref_wass_dist)-1)
                        
                        # pvalue for wass, 2 sided
                        # pvalue = 2 * stats.t.cdf(-np.abs(t_score), len(ref_wass_dist)-1)
                        # pvalue = 2 * stats.norm.cdf(-np.abs(t_score))

                        # print('doing shapiro')
                        # prev_shapiro_coeff, prev_shapiro_pval = stats.shapiro(prev_shuffled)
                        # curr_shapiro_coeff, curr_shapiro_pval = stats.shapiro(curr_shuffled)

                        regular_dict['signature'].append([prev_path, curr_path])
                        regular_dict['name'].append(name)
                        regular_dict['date'].append(date)
                        regular_dict['depth'].append(depth)
                        regular_dict['unit_id'].append(cell_label)
                        regular_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                        regular_dict['session_ids'].append([prev_key, curr_key])
                        regular_dict['whole_wass'].append(wass)
                        regular_dict['spike_density_wass'].append(spike_dens_wass)
                        regular_dict['z_score'].append(z_score)
                        regular_dict['mod_z_score'].append(mod_z_score)
                        regular_dict['p_value'].append(pvalue)
                        regular_dict['mod_p_value'].append(mod_pvalue)
                        # regular_dict['shapiro_pval'].append([prev_shapiro_pval, curr_shapiro_pval])
                        # regular_dict['shapiro_coeff'].append([prev_shapiro_coeff, curr_shapiro_coeff])
                        regular_dict['base_mean'].append(ref_wass_mean)
                        regular_dict['base_std'].append(ref_wass_std)
                        regular_dict['median'].append(median)
                        regular_dict['mad'].append(mad)

                        curr_fr_rate = len(curr_pts) / (curr_spike_pos_t[-1] - curr_spike_pos_t[0])
                        prev_fr_rate = len(prev_pts) / (prev_spike_pos_t[-1] - prev_spike_pos_t[0])
                        fr_rate_ratio = curr_fr_rate / prev_fr_rate
                        fr_rate_change = curr_fr_rate - prev_fr_rate

                        regular_dict['fr_rate'].append([prev_fr_rate, curr_fr_rate])
                        regular_dict['fr_rate_ratio'].append(fr_rate_ratio)
                        regular_dict['fr_rate_change'].append(fr_rate_change)

                        # regular_dict['total_fr_change'].append(np.nansum(curr_ratemap) - np.nansum(prev_ratemap))
                        # regular_dict['avg_fr_change'].append(np.nanmean(target_weights) - np.mean(source_weights))
                        # regular_dict['std_fr_change'].append(np.float64(np.std(target_weights) - np.std(source_weights)))

                        regular_dict['n_repeats'].append(settings['n_repeats'])
                        regular_dict['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])
                        regular_dict['cylinder'].append(cylinder)
                        assert prev.shape == curr.shape
                        regular_dict['ratemap_dims'].append(curr.shape)
                        if settings['downsample']:
                            regular_dict['downsample_factor'].append(settings['downsample_factor'])
                        else:
                            regular_dict['downsample_factor'].append(1)

                        if settings['plotRegular']:
                            plot_regular_remapping(prev, curr, regular_dict, data_dir)

                        if settings['plotShuffled']:
                            plot_shuffled_regular_remapping(prev_shuffled, curr_shuffled, ref_wass_dist, prev_shuffled_sample, curr_shuffled_sample, regular_dict, data_dir)

                        if settings['runFields']:

                            image_prev, n_labels_prev, source_labels, source_centroids, field_sizes_prev = blobs_dict[prev_id]

                            image_curr, n_labels_curr, target_labels, target_centroids, field_sizes_curr = blobs_dict[curr_id]

                            if len(np.unique(target_labels)) > 1 and len(np.unique(source_labels)) > 1:

                                """
                                cumulative_dict has field/centroid/binary wass, cumulative = all fields used
                                'field_wass' is EMD on ALL fields for norm/unnorm RATE remapping
                                'centroid_wass' is EMD on ALL field centroids for norm/unnorm LOCATION remapping (i.e. field centre points averaged + EMD calculated)
                                'binary_wass' is EMD on ALL fields (binary) for norm/unnorm LOCATION remapping (i.e. unweighted such that each pt contributes equally within field)

                                permute_dict has field/centroid/binary wass, permute = all combinations of single fields for given session pair
                                'field_wass' is EMD on SINGLE fields for norm/unnorm RATE remapping
                                'centroid_wass' is EMD on SINGLE field centroids for norm/unnorm LOCATION remapping (i.e. EMD calculated directly between diff centroid pairs across sessions)
                                'binary_wass' is EMD on SINGLE fields (binary) for norm/unnorm LOCATION remapping (i.e. unweighted such that each pt contributes equally within field)
                                """
                                permute_dict, cumulative_dict = compute_centroid_remapping(target_labels, source_labels, curr_spatial_spike_train, prev_spatial_spike_train, target_centroids, source_centroids, settings)

                                y, x = curr.shape
                                h, w = rate_map_obj.arena_size
                                bin_area = h/y * w/x
                                field_count = [len(np.unique(source_labels)) - 1,len(np.unique(target_labels)) - 1]

                                for centroid_score in settings['centroid_scores']:

                                    score_key = centroid_score + '_wass'

                                    centroid_dict['score'].append(centroid_score)
                                    # centroid_dict['signature'].append(curr_path)
                                    centroid_dict['name'].append(name)
                                    centroid_dict['date'].append(date)
                                    centroid_dict['depth'].append(depth)
                                    centroid_dict['bin_area'].append(bin_area[0])
                                    centroid_dict['field_count'].append(field_count)
                                    centroid_dict['unit_id'].append(cell_label)
                                    centroid_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                                    centroid_dict['session_ids'].append([prev_key, curr_key])
                                    centroid_dict['signature'].append([prev_path, curr_path])
                                    centroid_dict['arena_size'].append(curr_spatial_spike_train.arena_size)
                                    centroid_dict['cylinder'].append(cylinder)
                                    centroid_dict['ratemap_dims'].append(curr.shape)

                                    # field wass is weighed, centroid wass is centre pts, binary wass is unweighed (as if was bianry map with even weight so one pt per position in map)
                                    # centroid_dict['cumulative_wass'].append([cumulative_dict['field_wass'], cumulative_dict['centroid_wass'], cumulative_dict['binary_wass']])
                                    centroid_dict['cumulative_wass'].append(cumulative_dict[score_key])

                                    copy_labels = np.copy(source_labels)
                                    copy_labels[np.isnan(prev)] = 0
                                    copy_labels[copy_labels != 0] = 1
                                    val_r, val_c = np.where(copy_labels == 1)
                                    # cumulative_source_coverage = np.max(field_sizes_prev)
                                    cumulative_source_coverage = len(np.where(copy_labels != 0)[0])/len(np.where(~np.isnan(prev))[0])
                                    cumulative_source_area = len(np.where(copy_labels == 1)[0])
                                    cumulative_source_rate = np.sum(prev[val_r, val_c])

                                    copy_labels = np.copy(target_labels)
                                    copy_labels[np.isnan(curr)] = 0
                                    copy_labels[copy_labels != 0] = 1
                                    val_r, val_c = np.where(copy_labels == 1)
                                    # cumulative_target_coverage = np.max(field_sizes_curr)
                                    cumulative_target_coverage = len(np.where(copy_labels != 0)[0])/len(np.where(~np.isnan(curr))[0])
                                    cumulative_target_area = len(np.where(copy_labels == 1)[0])
                                    cumulative_target_rate = np.sum(curr[val_r, val_c])

                                    # cumulative_source_coverage = np.sum(field_sizes_prev)
                                    # cumulative_source_area = len(np.where(source_labels != 0)[0])
                                    # cumulative_source_rate = np.sum(prev[source_labels != 0])
                                    # cumulative_target_coverage = np.sum(field_sizes_curr)
                                    # cumulative_target_area = len(np.where(target_labels != 0)[0])
                                    # cumulative_target_rate = np.sum(curr[target_labels != 0])

                                    centroid_dict['cumulative_coverage'].append([cumulative_source_coverage, cumulative_target_coverage])
                                    centroid_dict['cumulative_area'].append([cumulative_source_area, cumulative_target_area])
                                    centroid_dict['cumulative_rate'].append([cumulative_source_rate, cumulative_target_rate])

                                                            # wass_args = [permute_dict['field_wass'], permute_dict['binary_wass'], permute_dict['centroid_wass']]
                                    wass_to_add = permute_dict[score_key]

                                    # need to make test_wass save as in fill centroid dict
                                    # centroid_dict['test_wass'].append(test_wass)
                                    # centroid_dict['centroid_wass'].append(centroid_wass)

                                    centroid_dict = _fill_centroid_dict(centroid_dict, max_centroid_count, wass_to_add, permute_dict['pairs'], permute_dict['coords'], prev, source_labels, field_sizes_prev, curr, target_labels, field_sizes_curr)

                                if settings['plotFields']:
                                    if cylinder:
                                        target_labels = flat_disk_mask(target_labels)
                                        source_labels = flat_disk_mask(source_labels)
                                    plot_fields_remapping(source_labels, target_labels, prev_spatial_spike_train, curr_spatial_spike_train, source_centroids, target_centroids, centroid_dict, data_dir, settings, cylinder=cylinder)

                        # remapping_indices[cell_label-1].append(i-1)

                        # remapping_session_ids[cell_label-1].append([i-1,i])

                        # c += 1

                    if prev is not None and settings['runTemporal']:

                        if prev_spike_times is None:
                            prev_spike_times = prev_spatial_spike_train.spike_times

                            # Shuffle spike trains within each row
                            # prev_shuffled_temporal = np.tile(prev_spikes, (num_shuffles, 1))
                            # prev_shuffled_indices = np.random.rand(num_shuffles, prev_shuffled_temporal.shape[1]).argsort(axis=1)
                            # prev_shuffled_temporal = np.take_along_axis(prev_shuffled_temporal, prev_shuffled_indices, axis=1)
                            # jitter_range = 0.05
                            # prev_shuffled_temporal = prev_spike_times + np.random.uniform(-jitter_range, jitter_range, size=len(spike_train))


                        curr_spike_times = curr_spatial_spike_train.spike_times

                        # Define the time bins to convert spike times to spike trains
                        # bin_width = 0.1  # Width of each time bin
                        # bin_edges = np.arange(0, np.max([np.max(prev_spike_times), np.max(curr_spike_times)]) + bin_width, bin_width)
                        # Calculate the pairwise distances between spikes
                        # distances = bin_edges[:-1] + bin_width / 2  
                        # distance_matrix = cdist(distances.reshape(-1, 1), distances.reshape(-1, 1))

                        # Perform shuffling and calculate EMD scores
                        num_shuffles = settings['n_temporal_shuffles']
                        # if prev_spikes is None:
                        #     prev_spikes = np.histogram(prev_spike_times, bin_edges)[0]
                        #     prev_spikes = prev_spikes / np.sum(prev_spikes)

                        # Convert spike times to spike trains with spike values
                        # curr_spikes = np.histogram(curr_spike_times, bin_edges)[0]
                        # curr_spikes = curr_spikes / np.sum(curr_spikes)

                        # curr_shuffled_temporal = np.tile(curr_spikes, (num_shuffles, 1))
                        # curr_shuffled_indices = np.random.rand(num_shuffles, curr_shuffled_temporal.shape[1]).argsort(axis=1)
                        # curr_shuffled_temporal = np.take_along_axis(curr_shuffled_temporal, curr_shuffled_indices, axis=1)
                        # print(prev_shuffled_temporal.shape, curr_shuffled_temporal.shape)
                        # print(prev_spikes.shape, curr_spikes.shape)
                        # print(prev_shuffled_indices.shape, curr_shuffled_indices.shape)
                        print('computing shuffled temporal emd')
                        # ref_emd_dist = compute_null_emd(prev_spike_times, curr_spike_times,num_shuffles)
                        # Calculate EMD scores for shuffled spike trains
                        # ref_emd_dist =[]
                        # for i in range(num_shuffles):
                        #     prev_shuffled_sample = prev_shuffled_temporal[i].flatten() / np.sum(prev_shuffled_temporal[i])
                        #     curr_shuffled_sample = curr_shuffled_temporal[i].flatten() / np.sum(curr_shuffled_temporal[i])
                        #     # print(prev_shuffled_sample.shape, curr_shuffled_sample.shape)
                        #     # ref_emd = emd(prev_shuffled_sample, curr_shuffled_sample, distance_matrix)
                        #     ref_emd = compute_emd(prev_shuffled_sample, curr_shuffled_sample)
                        #     ref_emd_dist.append(ref_emd)
                        # Compute the observed EMD score between the original spike trains
                        # observed_emd = emd(prev_spikes.squeeze(), curr_spikes.squeeze(), distance_matrix)
                        observed_emd = compute_emd(prev_spike_times, curr_spike_times)
                        # ref_emd_mean = np.mean(ref_emd_dist)
                        # ref_emd_std = np.std(ref_emd_dist)
                        # z_score = (observed_emd - ref_emd_mean) / (ref_emd_std)
                        print('doing modified z score')
                        # mod_z_score, median, mad = compute_modified_zscore(observed_emd, ref_emd_dist)
                        # assert len(ref_emd_dist) == settings['n_temporal_shuffles'], 'n_repeats does not match length of ref_emd_dist'

                        # pvalue = stats.t.cdf(z_score, len(ref_emd_dist)-1)
                        # mod_pvalue = stats.t.cdf(mod_z_score, len(ref_emd_dist)-1)
                        prev_duration = prev_spatial_spike_train.session_metadata.session_object.get_spike_data()['spike_cluster'].duration
                        curr_duration = curr_spatial_spike_train.session_metadata.session_object.get_spike_data()['spike_cluster'].duration

                        curr_fr_rate = len(curr_spike_times) / prev_duration
                        prev_fr_rate = len(prev_spike_times) / curr_duration
                        fr_rate_ratio = curr_fr_rate / prev_fr_rate
                        fr_rate_change = curr_fr_rate - prev_fr_rate

                        # # Calculate the Earth Mover's Distance (EMD) using the spike values and distances
                        # emd = emd_samples(prev_spikes, curr_spikes, distance_matrix)
                            
                        temporal_dict['signature'].append([prev_path, curr_path])
                        temporal_dict['depth'].append(depth)
                        temporal_dict['name'].append(name)
                        temporal_dict['date'].append(date)
                        temporal_dict['tetrode'].append(animal.animal_id.split('tet')[-1])
                        temporal_dict['unit_id'].append(cell_label)
                        temporal_dict['session_ids'].append([prev_key, curr_key])
                        temporal_dict['emd'].append(observed_emd)
                        # temporal_dict['z_score'].append(z_score)
                        # temporal_dict['p_value'].append(pvalue)
                        # temporal_dict['base_mean'].append(ref_emd_mean)a
                        # temporal_dict['base_std'].append(ref_emd_std)
                        # temporal_dict['mod_z_score'].append(mod_z_score)
                        # temporal_dict['mod_p_value'].append(mod_pvalue)
                        # temporal_dict['median'].append(median)
                        # temporal_dict['mad'].append(mad)
                        temporal_dict['fr_rate'].append([prev_fr_rate, curr_fr_rate])
                        temporal_dict['fr_rate_ratio'].append(fr_rate_ratio)
                        temporal_dict['fr_rate_change'].append(fr_rate_change)
                        # temporal_dict['n_repeats'].append(len(ref_emd_dist))
                        temporal_dict['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])


                    
                    prev = curr
                    prev_spikes = curr_spikes
                    prev_spike_times = curr_spike_times
                    prev_shuffled_temporal = curr_shuffled_temporal
                    prev_ratemap = curr_ratemap
                    prev_spatial_spike_train = curr_spatial_spike_train
                    prev_id = curr_id
                    prev_cell = curr_cell
                    prev_key = curr_key
                    prev_path = curr_path
                    prev_shuffled = curr_shuffled
                    # prev_plot = curr_plot
            
            # If there are context specific or otherwise specific groups to compare, can set those ids in settings
            # Will perform session to session remapping segregated by groups definned in settings
            if settings['runUniqueGroups']:
                # for category group (group of session ids)
                for categ in session_comp_categories:
                    categories = session_comp_categories[categ]
                    prev_key = None
                    prev = None 
                    prev_id = None
                    prev_spatial = None
                    prev_cell = None

                    # For session in that category
                    for ses in categories:
                        seskey = 'session_' + str(ses)
                        ses = animal.sessions[seskey]
                        path = ses.session_metadata.file_paths['tet']
                        fname = path.split('/')[-1].split('.')[0]
                        curr_id = str(animal.animal_id) + '_' + str(seskey) + '_' + str(cell.cluster.cluster_label)
                        
                        if settings['disk_arena']: 
                            cylinder = True
                        else:
                            cylinder, _ = check_disk_arena(fname)

                        ensemble = ses.get_cell_data()['cell_ensemble']

                        if cell_label in ensemble.get_cell_label_dict():
                            cell = ensemble.get_cell_by_id(cell_label)

                            # spatial_spike_train = cell.stats_dict['cell_stats']['spatial_spike_train']
                            spatial_spike_train = cell.stats_dict['spatial_spike_train'] 

                            rate_map_obj = spatial_spike_train.get_map('rate')

                            if settings['normalizeRate']:
                                rate_map, _ = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])
                            else:
                                _, rate_map = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])

                            assert rate_map.shape == (settings['ratemap_dims'][0], settings['ratemap_dims'][1]), 'Wrong ratemap shape {} vs settings shape {}'.format(rate_map.shape, (settings['ratemap_dims'][0], settings['ratemap_dims'][1]))
                            
                            # Disk mask ratemap
                            if cylinder:
                                curr = flat_disk_mask(rate_map)
                                if settings['downsample']:
                                    curr_ratemap = _downsample(rate_map, settings['downsample_factor'])
                                    curr_ratemap = flat_disk_mask(curr_ratemap)
                                else:
                                    curr_ratemap = curr
                                row, col = np.where(~np.isnan(curr_ratemap))
                                disk_ids = np.array([row, col]).T
                            else:
                                curr = rate_map
                                if settings['downsample']:
                                    curr_ratemap = _downsample(rate_map, settings['downsample_factor']) 
                                else:
                                    curr_ratemap = curr
                                disk_ids = None

                            curr_spatial = spatial_spike_train
                            curr_cell = cell
                            curr_key = seskey
                            curr_path = ses.session_metadata.file_paths['tet'].split('/')[-1].split('.')[0]

                            if prev is not None:
                                if prev_shuffled is None:  
                                    prev_shuffled = shuffled_ratemap_dict[prev_id]
                                curr_shuffled = shuffled_ratemap_dict[curr_id]
                                # get x and y pts for spikes in pair of sessions (prev and curr) for a given comparison

                                # prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial.get_spike_positions()
                                prev_spike_pos_x, prev_spike_pos_y, prev_spike_pos_t = prev_spatial_spike_train.spike_x, prev_spatial_spike_train.spike_y, prev_spatial_spike_train.new_spike_times
                                prev_pts = np.array([prev_spike_pos_x, prev_spike_pos_y]).T

                                # curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial.get_spike_positions()
                                curr_spike_pos_x, curr_spike_pos_y, curr_spike_pos_t = curr_spatial_spike_train.spike_x, curr_spatial_spike_train.spike_y, curr_spatial_spike_train.new_spike_times
                                curr_pts = np.array([curr_spike_pos_x, curr_spike_pos_y]).T

                                if settings['rotate_evening']:
                                    if 'evening' or 'rotated' in curr_path.lower():
                                        prev_ratemap = ndimage.rotate(prev_ratemap, settings['rotate_angle'])
                                        curr_ratemap = ndimage.rotate(curr_ratemap, settings['rotate_angle'])

                                        if settings['rotate_angle'] == 90:
                                            prev_pts = np.array([prev_pts[:,1], -prev_pts[:,0]]).T
                                            curr_pts = np.array([curr_pts[:,1], -curr_pts[:,0]]).T
                                            print('rotating 90 degrees for {}'.format(curr_path))
                                        else:
                                            raise ValueError('Rotation angle not supported {}'.format(settings['rotate_angle']))

                                y, x = prev_ratemap.shape
                                # find indices of not nan 
                                row_prev, col_prev = np.where(~np.isnan(prev_ratemap))
                                row_curr, col_curr = np.where(~np.isnan(curr_ratemap))

                                # # for first map
                                # if prev_shuffled is None:  
                                #     prev_shuffled_samples = list(map(lambda x: _single_shuffled_sample(prev_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                                #     if cylinder:
                                #         prev_shuffled_samples = list(map(lambda x: flat_disk_mask(x), prev_shuffled_samples))
                                #     prev_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_prev, col_prev))), prev_shuffled_samples))

                                # curr_shuffled_samples = list(map(lambda x: _single_shuffled_sample(curr_spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                                # if cylinder:
                                #     curr_shuffled_samples = list(map(lambda x: flat_disk_mask(x), curr_shuffled_samples))   
                                # curr_shuffled = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row_curr, col_curr))), curr_shuffled_samples))

                                # assert row_prev.all() == row_curr.all() and col_prev.all() == col_curr.all(), 'Nans in different places'

                                height_bucket_midpoints, width_bucket_midpoints = _get_ratemap_bucket_midpoints(prev_spatial_spike_train.arena_size, y, x)
                                height_bucket_midpoints = height_bucket_midpoints[row_curr]
                                width_bucket_midpoints = width_bucket_midpoints[col_curr]
                                source_weights = np.array(list(map(lambda x, y: prev_ratemap[x,y], row_prev, col_prev)))
                                target_weights = np.array(list(map(lambda x, y: curr_ratemap[x,y], row_curr, col_curr)))
                                source_weights = source_weights / np.sum(source_weights)
                                target_weights = target_weights / np.sum(target_weights)
                                coord_buckets_curr = np.array(list(map(lambda x, y: [height_bucket_midpoints[x],width_bucket_midpoints[y]], row_curr, col_curr)))
                                coord_buckets_prev = np.array(list(map(lambda x, y: [height_bucket_midpoints[x],width_bucket_midpoints[y]], row_prev, col_prev)))

                                spike_dens_wass = pot_sliced_wasserstein(prev_pts, curr_pts, n_projections=settings['n_projections'])
                                    # elif rate_score == 'whole':
                                        # This is EMD on whole map for normalized/unnormalized rate remapping
                                wass = pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, source_weights, target_weights, n_projections=settings['n_projections'])
                                ref_wass_dist = list(map(lambda x, y: pot_sliced_wasserstein(coord_buckets_prev, coord_buckets_curr, x/np.sum(x), y/np.sum(y), n_projections=settings['n_shuffle_projections']), prev_shuffled, curr_shuffled))
                                ref_wass_mean = np.mean(ref_wass_dist)
                                ref_wass_std = np.std(ref_wass_dist)
                                z_score = (wass - ref_wass_mean) / (ref_wass_std)
                                mod_z_score, median, mad = compute_modified_zscore(wass, ref_wass_dist)

                                assert len(ref_wass_dist) == settings['n_repeats'], 'n_repeats does not match length of ref_wass_dist'
        
                                pvalue = stats.t.cdf(z_score, len(ref_wass_dist)-1)
                                mod_pvalue = stats.t.cdf(mod_z_score, len(ref_wass_dist)-1)
                                # pvalue for wass, 2 sided
                                # pvalue = 2 * stats.t.cdf(-np.abs(t_score), len(ref_wass_dist)-1)
                                # pvalue = 2 * stats.norm.cdf(-np.abs(t_score))

                                # prev_shapiro_coeff, prev_shapiro_pval = stats.shapiro(prev_shuffled)
                                # curr_shapiro_coeff, curr_shapiro_pval = stats.shapiro(curr_shuffled)

                                context_dict[categ]['signature'].append([prev_path, curr_path])
                                context_dict[categ]['name'].append(name)
                                context_dict[categ]['date'].append(date)
                                context_dict[categ]['depth'].append(depth)
                                context_dict[categ]['unit_id'].append(cell_label)
                                context_dict[categ]['tetrode'].append(animal.animal_id.split('tet')[-1])
                                context_dict[categ]['session_ids'].append([prev_key, curr_key])
                                context_dict[categ]['whole_wass'].append(wass)
                                context_dict[categ]['spike_density_wass'].append(spike_dens_wass)
                                context_dict[categ]['z_score'].append(z_score)
                                context_dict[categ]['mod_z_score'].append(mod_z_score)
                                context_dict[categ]['p_value'].append(pvalue)
                                context_dict[categ]['mod_p_value'].append(mod_pvalue)
                                # context_dict[categ]['shapiro_pval'].append([prev_shapiro_pval, curr_shapiro_pval])
                                # context_dict[categ]['shapiro_coeff'].append([prev_shapiro_coeff, curr_shapiro_coeff])
                                context_dict[categ]['base_mean'].append(ref_wass_mean)
                                context_dict[categ]['base_std'].append(ref_wass_std)
                                context_dict[categ]['median'].append(median)
                                context_dict[categ]['mad'].append(mad)

                                curr_fr_rate = len(curr_pts) / (curr_spike_pos_t[-1] - curr_spike_pos_t[0])
                                prev_fr_rate = len(prev_pts) / (prev_spike_pos_t[-1] - prev_spike_pos_t[0])
                                fr_rate_ratio = curr_fr_rate / prev_fr_rate
                                fr_rate_change = curr_fr_rate - prev_fr_rate

                                context_dict[categ]['fr_rate'].append([prev_fr_rate, curr_fr_rate])
                                context_dict[categ]['fr_rate_ratio'].append(fr_rate_ratio)
                                context_dict[categ]['fr_rate_change'].append(fr_rate_change)

                                context_dict[categ]['n_repeats'].append(settings['n_repeats'])
                                context_dict[categ]['arena_size'].append([prev_spatial_spike_train.arena_size, curr_spatial_spike_train.arena_size])
                                context_dict[categ]['cylinder'].append(cylinder)
                                assert prev.shape == curr.shape
                                context_dict[categ]['ratemap_dims'].append(curr.shape)
                                if settings['downsample']:
                                    context_dict[categ]['downsample_factor'].append(settings['downsample_factor'])
                                else:
                                    context_dict[categ]['downsample_factor'].append(1)
                                
                            prev = curr
                            prev_cell = curr_cell
                            prev_spatial = curr_spatial
                            prev_key = curr_key
                            prev_path = curr_path

            if settings['plotMatchedWaveforms'] and settings['runRegular']:     
                plot_matched_sesssion_waveforms(cell_session_appearances, settings, regular_dict, data_dir)
            # c += 1

            to_save = {'regular': regular_dict, 'object': obj_dict, 'centroid': centroid_dict, 'context': context_dict, 'temporal': temporal_dict}
            if settings['runRegular']:
                df = pd.DataFrame(to_save['regular'])
                # df.to_csv(PROJECT_PATH + '/_prototypes/cell_remapping/remapping_output' + '/rate_remapping.csv')
                # check if file exists
                if isStart:
                    path = data_dir + '/remapping_output/regular_remapping.xlsx'
                    # check if file exists, otherwise check file2, otherwise file3, etc... Once not found then create new file
                    if not os.path.isfile(path):
                        regular_path_to_use = path
                    else:
                        counter = 2
                        # TO DO
                        # save to csv iteratively - DONE?
                        # rewrite get_spike_positions
                        # remove shapiro - DONE
                        # store spike positions instead of recomputing - DONE
                        # store shuffled dists for context remapping - DONE
                        # use full opexebo shuffle fxn with iterations = 1000 and return that - DONE
                        # check if value between 10**2 and 10**3 for true wass dist
                        # check of value 50 as good as 10**2 for shuffled dist
                        # test 16,16 ratemaps since we doing whole map reg remapping not centroids, make sure compare wass and ref wass
                        while os.path.isfile(data_dir + '/remapping_output/regular_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        regular_path_to_use = data_dir + '/remapping_output/regular_remapping_' + str(counter) + '.xlsx'
                    # create workbook 
                    writer = pd.ExcelWriter(regular_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    # writer.flush()
                    writer.close()

                else:
                    book = load_workbook(regular_path_to_use)
                    writer = pd.ExcelWriter(regular_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    # writer.flush()
                    writer.close()
                    book.save(regular_path_to_use)
                    # book.close()

            if settings['hasObject']:
                df = pd.DataFrame(to_save['object'])
                if isStart:
                    path = data_dir + '/remapping_output/obj_remapping.xlsx'
                    if not os.path.isfile(path):
                        obj_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/obj_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        obj_path_to_use = data_dir + '/remapping_output/obj_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(obj_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(obj_path_to_use)
                    writer = pd.ExcelWriter(obj_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    writer.close()
                    book.save(obj_path_to_use)
                    # book.close()
            
            if settings['runFields']:
                df = pd.DataFrame(to_save['centroid'])
                if isStart:
                    path = data_dir + '/remapping_output/centroid_remapping.xlsx'
                    if not os.path.isfile(path):
                        centroid_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/centroid_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        centroid_path_to_use = data_dir + '/remapping_output/centroid_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(centroid_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(centroid_path_to_use)
                    writer = pd.ExcelWriter(centroid_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    writer.close()
                    book.save(centroid_path_to_use)
            if settings['runUniqueGroups']:
        
                for context in to_save['context']:
                    df = pd.DataFrame(to_save['context'][context])
                    if isStart:
                        path = data_dir + '/remapping_output/context_' + context + '_remapping.xlsx'
                        if not os.path.isfile(path):
                            context_path_to_use = path
                        else:
                            counter = 2
                            while os.path.isfile(data_dir + '/remapping_output/context_' + context + '_' + str(counter) + '.xlsx'):
                                counter += 1
                            context_path_to_use = data_dir + '/remapping_output/context_' + context + '_' + str(counter) + '.xlsx'
                        writer = pd.ExcelWriter(context_path_to_use, engine='openpyxl')
                        df.to_excel(writer, sheet_name='Summary')
                        writer.save()
                        writer.close()
                        context_paths[context] = context_path_to_use
                    else:
                        context_path_to_use = context_paths[context]
                        book = load_workbook(context_path_to_use)
                        writer = pd.ExcelWriter(context_path_to_use, engine='openpyxl')
                        writer.book = book
                        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                        df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                        # writer.save()
                        writer.close()
                        book.save(context_path_to_use)
            if settings['runTemporal']:
                df = pd.DataFrame(to_save['temporal'])
                if isStart:
                    path = data_dir + '/remapping_output/temporal_remapping.xlsx'
                    if not os.path.isfile(path):
                        temporal_path_to_use = path
                    else:
                        counter = 2
                        while os.path.isfile(data_dir + '/remapping_output/temporal_remapping_' + str(counter) + '.xlsx'):
                            counter += 1
                        temporal_path_to_use = data_dir + '/remapping_output/temporal_remapping_' + str(counter) + '.xlsx'
                    writer = pd.ExcelWriter(temporal_path_to_use, engine='openpyxl')
                    df.to_excel(writer, sheet_name='Summary')
                    writer.save()
                    writer.close()
                else:
                    book = load_workbook(temporal_path_to_use)
                    writer = pd.ExcelWriter(temporal_path_to_use, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Summary', header=False, startrow=writer.sheets['Summary'].max_row)
                    # writer.save()
                    writer.close()
                    book.save(temporal_path_to_use)
            isStart = False

    # return {'regular': regular_dict, 'object': obj_dict, 'centroid': centroid_dict, 'context': dict}

def compute_null_emd(spike_train_a, spike_train_b, num_iterations):
    combined_spike_train = np.concatenate([spike_train_a, spike_train_b])
    num_spikes = len(spike_train_a)
    num_total_spikes = len(combined_spike_train)

    # Repeat the combined spike train for the desired number of iterations
    repeated_spike_train = np.tile(combined_spike_train, (num_iterations, 1))

    # Shuffle the spike trains along the second axis (columns)
    np.apply_along_axis(np.random.shuffle, axis=1, arr=repeated_spike_train)

    # Extract the shuffled spike trains for A and B
    shuffled_spike_train_a = repeated_spike_train[:, :num_spikes]
    shuffled_spike_train_b = repeated_spike_train[:, num_spikes:num_total_spikes]

    emd_values = np.empty(num_iterations)
    for i in range(num_iterations):
        emd_values[i] = compute_emd(shuffled_spike_train_a[i], shuffled_spike_train_b[i])

    return emd_values

# Perform further analysis or calculations with the 'emd_values' array


def compute_emd(spike_train_a, spike_train_b):
    # Determine the start and end times for aligning the spike trains
    start_time = np.min([np.min(spike_train_a), np.min(spike_train_b)])
    end_time = np.max([np.max(spike_train_a), np.max(spike_train_b)])

    bins = np.arange(start_time, end_time + 1)
    # Create aligned spike trains
    aligned_a, _ = np.histogram(spike_train_a, bins=bins)
    aligned_b, _ = np.histogram(spike_train_b, bins=bins)

    # Compute the cumulative distribution functions (CDFs)
    cdf_a = np.cumsum(aligned_a) / len(spike_train_a)
    cdf_b = np.cumsum(aligned_b) / len(spike_train_b)

    # Compute the EMD by integrating the absolute difference between CDFs
    emd = np.sum(np.abs(cdf_a - cdf_b))

    return emd

def _fill_cell_type_stats(inp_dict, prev_cell, curr_cell):
    inp_dict['information'].append([prev_cell.stats_dict['cell_stats']['spatial_information_content'],curr_cell.stats_dict['cell_stats']['spatial_information_content'],curr_cell.stats_dict['cell_stats']['spatial_information_content']-prev_cell.stats_dict['cell_stats']['spatial_information_content']])
    inp_dict['grid_score'].append([prev_cell.stats_dict['cell_stats']['grid_score'],curr_cell.stats_dict['cell_stats']['grid_score'],curr_cell.stats_dict['cell_stats']['grid_score']-prev_cell.stats_dict['cell_stats']['grid_score']])
    inp_dict['b_top'].append([prev_cell.stats_dict['cell_stats']['b_score_top'],curr_cell.stats_dict['cell_stats']['b_score_top'],curr_cell.stats_dict['cell_stats']['b_score_top']-prev_cell.stats_dict['cell_stats']['b_score_top']])
    inp_dict['b_bottom'].append([prev_cell.stats_dict['cell_stats']['b_score_bottom'],curr_cell.stats_dict['cell_stats']['b_score_bottom'],curr_cell.stats_dict['cell_stats']['b_score_bottom']-prev_cell.stats_dict['cell_stats']['b_score_bottom']])
    inp_dict['b_right'].append([prev_cell.stats_dict['cell_stats']['b_score_right'],curr_cell.stats_dict['cell_stats']['b_score_right'],curr_cell.stats_dict['cell_stats']['b_score_right']-prev_cell.stats_dict['cell_stats']['b_score_right']])
    inp_dict['b_left'].append([prev_cell.stats_dict['cell_stats']['b_score_left'],curr_cell.stats_dict['cell_stats']['b_score_left'],curr_cell.stats_dict['cell_stats']['b_score_left']-prev_cell.stats_dict['cell_stats']['b_score_left']])
                            
    # inp_dict['speed_score'].append([prev_cell.stats_dict['cell_stats']['speed_score'],curr_cell.stats_dict['cell_stats']['speed_score'],curr_cell.stats_dict['cell_stats']['speed_score']-prev_cell.stats_dict['cell_stats']['speed_score']])
    # inp_dict['hd_score'].append([prev_cell.stats_dict['cell_stats']['hd_score'],curr_cell.stats_dict['cell_stats']['hd_score'],curr_cell.stats_dict['cell_stats']['hd_score']-prev_cell.stats_dict['cell_stats']['hd_score']])

    return inp_dict

def _read_location_from_file(path, cylinder, true_var):

    if not cylinder:
        object_location = path.split('/')[-1].split('-')[3].split('.')[0]
    else:
        items = path.split('/')[-1].split('-')
        idx = items.index(str(true_var)) + 2 # the object location is always 2 positions away from word denoting arena hape (e.g round/cylinder) defined by true_var
        # e.g. ROUND-3050-90_2.clu
        object_location = items[idx].split('.')[0].split('_')[0].lower()

    object_present = True
    if str(object_location) == 'no':
        object_present == False
        object_location = 'no'
    elif str(object_location) == 'zero':
        object_location = 0
    else:
        object_location = int(object_location)
        assert int(object_location) in [0,90,180,270], 'Failed bcs obj location is ' + str(int(object_location)) + ' and that is not in [0,90,180,270]'

    return object_location

def _aggregate_cell_info(animal, settings):
    ratemap_size = settings['ratemap_dims'][0]
    max_centroid_count = 0
    blobs_dict = {}
    # type_dict = {}
    # spatial_obj_dict = {}
    shuffled_ratemap_dict = {}
    shuffled_sample_dict = {}
    # for animal in study.animals:
    for onlyone in range(1):
        # get largest possible cell id
        # max_matched_cell_count = len(animal.sessions[sorted(list(animal.sessions.keys()))[-1]].get_cell_data()['cell_ensemble'].cells)
        max_matched_cell_count = max(list(map(lambda x: max(animal.sessions[x].get_cell_data()['cell_ensemble'].get_label_ids()), animal.sessions)))
        for k in range(int(max_matched_cell_count)):
            cell_label = k + 1
            prev_field_size_len = None 
            for i in range(len(list(animal.sessions.keys()))):
                seskey = 'session_' + str(i+1)
                ses = animal.sessions[seskey]
                ensemble = ses.get_cell_data()['cell_ensemble']
                pos_obj = ses.get_position_data()['position']

                path = ses.session_metadata.file_paths['tet']
                fname = path.split('/')[-1].split('.')[0]

                # Check if cylinder
                if settings['disk_arena']: 
                    cylinder = True
                else:
                    cylinder, _ = check_disk_arena(fname)

                if cell_label in ensemble.get_cell_label_dict():
                    cell = ensemble.get_cell_by_id(cell_label)
                    # if 'spatial_spike_train' in cell.stats_dict['cell_stats']:
                    #     spatial_spike_train = cell.stats_dict['cell_stats']['spatial_spike_train']
                    # else:
                    #     spatial_spike_train = ses.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': ses.get_position_data()['position']})
                    
                    spatial_spike_train = ses.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj})
                    cell.stats_dict['spatial_spike_train'] = spatial_spike_train

                    rate_map_obj = spatial_spike_train.get_map('rate')
                    
                    if settings['normalizeRate']:
                        rate_map, _ = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])
                    else:
                        _, rate_map = rate_map_obj.get_rate_map(new_size = settings['ratemap_dims'][0])

                    assert rate_map.shape == (settings['ratemap_dims'][0], settings['ratemap_dims'][1]), 'Wrong ratemap shape {} vs settings shape {}'.format(rate_map.shape, (settings['ratemap_dims'][0], settings['ratemap_dims'][1]))
                    
                    if settings['downsample']:
                        rate_map = _downsample(rate_map, settings['downsample_factor'])
                    if cylinder:
                        rate_map = flat_disk_mask(rate_map)

                    id = str(animal.animal_id) + '_' + str(seskey) + '_' + str(cell.cluster.cluster_label)
                    assert id not in blobs_dict, 'Duplicate cell id ' + str(id)
                        
                    if settings['hasObject'] or settings['runFields']:
                        image, n_labels, labels, centroids, field_sizes = map_blobs(spatial_spike_train, ratemap_size=ratemap_size, cylinder=cylinder, downsample=settings['downsample'], downsample_factor=settings['downsample_factor'])

                        labels, centroids, field_sizes = _sort_filter_centroids_by_field_size(rate_map, field_sizes, labels, centroids, spatial_spike_train.arena_size)

                        blobs_dict[id] = [image, n_labels, labels, centroids, field_sizes]

                        if prev_field_size_len is not None:
                            max_centroid_count = max(max_centroid_count, len(field_sizes) * prev_field_size_len)
                        else:
                            prev_field_size_len = len(field_sizes)

                    if settings['runRegular'] or settings['runUniqueGroups']:
                        row, col = np.where(~np.isnan(rate_map))
                        print('drawing shuffled samples')
                        # shuffled_samples = list(map(lambda x: _single_shuffled_sample(spatial_spike_train, settings), np.arange(settings['n_repeats'])))
                        norm, raw = spatial_spike_train.get_map('rate').get_rate_map(new_size = settings['ratemap_dims'][0], shuffle=True, n_repeats=settings['n_repeats'])
                        print(norm.shape, raw.shape, rate_map.shape)
                        shuffled_samples = list(map(lambda x, y: _single_shuffled_sample(x, y, settings), norm, raw))
                        print(np.array(shuffled_samples).shape)
                        if cylinder:
                            shuffled_samples = list(map(lambda x: flat_disk_mask(x), shuffled_samples))
                        print('turning into valid weights')
                        shuffled_ratemaps = list(map(lambda sample: np.array(list(map(lambda x, y: sample[x,y], row, col))), shuffled_samples))

                        shuffled_ratemap_dict[id] = shuffled_ratemaps
                        shuffled_sample_dict[id] = shuffled_samples[np.random.randint(0, settings['n_repeats'])]

                    # spatial_obj_dict[id] = spatial_spike_train

                    # type_dict[id] = []
                

    return max_centroid_count, blobs_dict, shuffled_ratemap_dict, shuffled_sample_dict


# def _sort_filter_centroids_by_field_size(prev, curr, field_sizes_source, field_sizes_target, blobs_map_source, blobs_map_target, centroids_prev, centroids_curr, arena_size):
def _sort_filter_centroids_by_field_size(rate_map, field_sizes, blobs_map, centroids, arena_size):
    height, width = arena_size
    y, x = rate_map.shape
    heightStep = height/y
    widthStep = width/x

    bin_area = heightStep * widthStep

    if type(bin_area) == list:
        assert len(bin_area) == 1
        bin_area = bin_area[0]

    # find not nan in prev/curr 
    row, col = np.where(np.isnan(rate_map))
    blobs_map[row, col] = 0

    # find bins for each label
    pks = []
    avgs = []
    lbls = []
    for k in np.unique(blobs_map):
        if k != 0:
            row, col = np.where(blobs_map == k)
            
            field = rate_map[row, col]

            # pk_rate = np.max(field)

            # get rate at centroid 
            c_row, c_col = centroids[k-1]
            # convert from decimals to bin by rounding to nearest int
            c_row = int(np.round(c_row))
            c_col = int(np.round(c_col))

            # pk_rate = rate_map[c_row, c_col]  
            # pk_rate = np.max(field)

            avg_rate = np.mean(field)
            pk_rate = np.sum(field)

            pks.append(pk_rate)
            avgs.append(avg_rate)
            lbls.append(k-1)
    
    pks = np.array(pks)
    ids = np.argsort(-pks)
    sort_idx = np.array(lbls)[ids]

    source_labels = np.zeros(blobs_map.shape)
    map_dict = {}
    # source_centroids = []
    # source_field_sizes = []
    all_filtered_out = True
    largest_label_id = None
    largest_field_area = 0
    for k in np.unique(blobs_map):
        if k != 0:
            row, col = np.where(blobs_map == k)
            field_area = (len(row) + len(col)) * bin_area
            if field_area > largest_field_area:
                    largest_label_id = k
                    largest_field_area = field_area

            if field_area > 22.5:
                idx_to_move_to = np.where(sort_idx == k-1)[0][0]
                # map_dict[k] = sort_idx_source[k-1] + 1
                map_dict[k] = idx_to_move_to + 1
                # source_centroids.append(centroids[k-1])
                all_filtered_out = False
            else:
                print('Blob filtered out with size less than 22.5 cm^2')
                map_dict[k] = 0
                # remove from sort_idx
                sort_idx = sort_idx[np.where(sort_idx != k-1)[0]]
        else:
            map_dict[k] = 0
    source_labels = np.vectorize(map_dict.get)(blobs_map)

    if len(sort_idx) > 0:
        source_centroids = np.asarray(centroids)[sort_idx]
        source_field_sizes = np.asarray(field_sizes)[sort_idx]
    else:
        source_centroids = np.asarray(centroids)
        source_field_sizes = np.asarray(field_sizes)

    if all_filtered_out:
        assert largest_label_id is not None
        map_dict[largest_label_id] = 1
        source_labels = np.vectorize(map_dict.get)(blobs_map)
        source_centroids = [np.asarray(centroids)[largest_label_id-1]]
        source_field_sizes = [np.asarray(field_sizes)[largest_label_id-1]]

    # print('blobshere')
    # print('all_filtered_out: ' + str(all_filtered_out))
    # print(pks, ids, lbls, sort_idx)
    # print(np.unique(source_labels), len(source_centroids), len(source_field_sizes))
    # print(map_dict, np.unique(blobs_map))
    assert len(np.unique(source_labels)) - 1 == len(source_centroids) == len(source_field_sizes)

    return source_labels, source_centroids, source_field_sizes

def _fill_centroid_dict(centroid_dict, max_centroid_count, wass_to_add, centroid_pairs, centroid_coords, source_map, source_labels, source_field_sizes, target_map, target_labels, target_field_sizes):
    # field_wass, centroid_wass, binary_wass = wass_args
    # to_add = wass_args
    for n in range(max_centroid_count):
        wass_key = 'c_wass_'+str(n+1)
        id_key = 'c_ids_'+str(n+1)
        vector_key = 'c_vector_'+str(n+1)
        coverage_key = 'c_coverage_'+str(n+1)
        area_key = 'c_area_'+str(n+1)
        rate_key = 'c_rate_'+str(n+1)

        if wass_key not in centroid_dict:
            centroid_dict[wass_key] = []
            centroid_dict[id_key] = []
            centroid_dict[coverage_key] = []
            centroid_dict[area_key] = []
            centroid_dict[rate_key] = []
            centroid_dict[vector_key] = []

        if n < len(wass_to_add):
            centroid_dict[wass_key].append(wass_to_add[n])
            centroid_dict[id_key].append(centroid_pairs[n])

            labels_curr = np.copy(source_labels)
            labels_curr[np.isnan(source_map)] = 0
            val_r, val_c = np.where(labels_curr == centroid_pairs[n][0])
            # take only field label = 1 = largest field = main
            # source_field_coverage = source_field_sizes[centroid_pairs[n][0]-1]
            source_field_coverage = len(np.where(labels_curr == centroid_pairs[n][0])[0])/len(np.where(~np.isnan(source_labels))[0])
            source_field_area = len(np.where(labels_curr == centroid_pairs[n][0])[0])
            source_field_rate = np.sum(source_map[val_r, val_c])

            labels_curr = np.copy(target_labels)
            labels_curr[np.isnan(target_map)] = 0
            val_r, val_c = np.where(labels_curr == centroid_pairs[n][1])
            # take only field label = 1 = largest field = main
            # target_field_coverage = target_field_sizes[centroid_pairs[n][1]-1]
            target_field_coverage = len(np.where(labels_curr == centroid_pairs[n][1])[0])/len(np.where(~np.isnan(target_labels))[0])
            target_field_area = len(np.where(labels_curr == centroid_pairs[n][1])[0])
            target_field_rate = np.sum(target_map[val_r, val_c])

            source_r, source_c = centroid_coords[n][0]
            target_r, target_c = centroid_coords[n][1]
            mag = np.linalg.norm(np.array((source_r, source_c)) - np.array((target_r, target_c)))
            pt1 = (source_r, source_c)
            pt2 = (target_r, target_c)
            angle = np.degrees(np.math.atan2(np.linalg.det([pt1,pt2]),np.dot(pt1,pt2)))

            
            centroid_dict[coverage_key].append([source_field_coverage, target_field_coverage])
            centroid_dict[area_key].append([source_field_area, target_field_area])
            centroid_dict[rate_key].append([source_field_rate, target_field_rate])
            centroid_dict[vector_key].append([pt1, pt2, mag, angle])
            
        else:
            centroid_dict[wass_key].append(np.nan)
            centroid_dict[id_key].append([np.nan])
            centroid_dict[coverage_key].append([np.nan, np.nan])
            centroid_dict[area_key].append([np.nan, np.nan])
            centroid_dict[rate_key].append([np.nan, np.nan])
            centroid_dict[vector_key].append([np.nan, np.nan, np.nan, np.nan])

        # if wass_key not in centroid_dict:
        #     centroid_dict[wass_key] = []
        #     centroid_dict[id_key] = []

        # if n < len(centroid_wass):
        #     centroid_dict[wass_key].append([field_wass[n], centroid_wass[n], binary_wass[n]])
        #     centroid_dict[id_key].append(centroid_pairs[n])
        # else:
        #     centroid_dict[wass_key].append([0,0,0])
        #     centroid_dict[id_key].append([0,0])

    return centroid_dict