import os
import sys
import traceback

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from _prototypes.unit_matcher.waveform import time_index, troughs
import warnings
from library.study_space import Session, Study, Animal
from scripts.batch_map.LEC_naming import LEC_naming_format, extract_name_lec
from _prototypes.cell_remapping.src.MEC_naming import MEC_naming_format, extract_name_mec
import tkinter as tk
from tkinter import filedialog
import time
from library.hafting_spatial_maps import SpatialSpikeTrain2D
from library.scores import rate_map_stats, rate_map_coherence
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from library.maps.map_utils import disk_mask
from PIL import Image
import numpy as np
from matplotlib import cm
import matplotlib.pyplot as plt
#import random
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension
from matplotlib import cm
# from opexebo.analysis import rate_map_stats, speed_score, rate_map_coherence
# Set matplotlib backend to Agg to prevent figures from popping up.
from matplotlib import use
from library.maps.firing_rate_vs_time import firing_rate_vs_time
from library.maps.filter_pos_by_speed import filter_pos_by_speed
from library.map_utils import _speed2D
from x_io.rw.axona.batch_read import make_study
import pandas as pd
import re

""" SETTINGS AT THE BOTTOM OF FILE """

def _check_single_format(filename, format, fxn):
    print(filename, format, fxn)
    if re.match(str(format), str(filename)) is not None:
        return fxn(filename)


def batch_map(study: Study, settings_dict: dict, saveDir=None, sum_sheet_count=None):
    """
    Computes rate maps across all animals, sessions, cells in a study.

    Use tasks dictionary as true/false flag with variable to compute
    e.g. {'rate_map': True, 'binary_map': False}
    """

    tasks = settings_dict['tasks']
    plotTasks = settings_dict['plotTasks']
    csv_header = settings_dict['header']

    # Root path creation
    save_dir = saveDir
    root_path = os.path.join(save_dir, 'Spatial_Shuffle_')
    run_number = 1
    while os.path.isdir(root_path+str(run_number)):
        run_number+=1

    root_path += str(run_number)
    os.mkdir(root_path)

    # Kernel size
    kernlen = int(settings_dict['smoothing_factor']*8)
    # Standard deviation size
    std = int(0.2*kernlen)

    # Set flag if no arguments were provided for later
    all_cells = False
    if (settings_dict['start_cell'] == settings_dict['end_cell'] == None):
        all_cells = True

    # Grabs headers whose value is true
    headers = [k for k, v in csv_header.items() if v]
    if 'border_score' in headers:
        idx = headers.index('border_score') 
        # print(idx)
        headers.remove('border_score')
        headers.insert(idx, 'border_score_top')
        headers.insert(idx+1, 'border_score_bottom')
        headers.insert(idx+2, 'border_score_left')
        headers.insert(idx+3, 'border_score_right')
        # print(headers)
        # stop()

    headers_dict = dict()

    for i, header in enumerate(headers):
        headers_dict[header] = get_column_letter(i+4)

    file_name = os.path.join(root_path, "spatial_shuffle_parameters.txt")
    with open(file_name, 'w') as f:
        for ky in settings_dict:
            f.write(str(ky) + ' is: ')
            f.write(str(settings_dict[ky]) + '\n')
        f.close()

    if study.animals is None:
        study.make_animals()
        print('Animals made, batching map')

    per_animal_tracker = 1
    per_animal_tetrode_tracker = 1

    animal_tet_count = {}
    animal_max_tet_count = {}
    animal_workbooks = {}
    sorted_animal_ids = np.unique(np.sort(study.animal_ids))

    one_for_parent_wb = xl.Workbook()
    sum_sheet = one_for_parent_wb['Sheet']
    sum_sheet.title = 'Summary'
    sum_sheet['A' + str(1)] = 'Session'
    sum_sheet['B' + str(1)] = 'Tetrode'
    sum_sheet['C' + str(1)] = 'Cell ID'

    for animalID in sorted_animal_ids:
        animal = study.get_animal_by_id(animalID)

        animal_id = animal.animal_id.split('_tet')[0]
        if animal_id not in animal_tet_count:
            animal_tet_count[animal_id] = 1
            # animal_sessions_tets_events[animal_id] = {}
            wb = xl.Workbook()
            animal_workbooks[animal_id] = wb
            animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
            sum_sheet = wb['Sheet']
            sum_sheet.title = 'Summary'
            if settings_dict['saveMethod'] == 'one_per_session':
                # animal_tet_count[animal_id] = np.ones(len(animal.sessions))
                animal_workbooks[animal_id] = {}
                for ses_key in animal.sessions:
                    wb = xl.Workbook()
                    sum_sheet = wb['Sheet']
                    sum_sheet.title = 'Summary'
                    animal_workbooks[animal_id][ses_key] = wb
            elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                # wb.remove_sheet('Summary') 
                pass
            elif settings_dict['saveMethod'] == 'one_for_parent':
                # Copy headers into excel file
                #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                wb = one_for_parent_wb
                sum_sheet = wb['Summary']
                # sum_sheet['A' + str(1)] = 'Session'
                # sum_sheet['B' + str(1)] = 'Tetrode'
                # sum_sheet['C' + str(1)] = 'Cell'
                for header in headers:
                    #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                    sum_sheet[headers_dict[header] + str(1)] = header
        else:
            if settings_dict['saveMethod'] != 'one_for_parent':
                animal_tet_count[animal_id] += 1
                # animal_tets[animal_id].append(animal)
                wb = animal_workbooks[animal_id]
            else:
                wb = one_for_parent_wb

        k = 1

        # col_count = 0 

        for session_key in animal.sessions:
            session = animal.sessions[session_key]

            c = 0

            pos_obj = session.get_position_data()['position']

            # excel_cell_index = 1
            per_session_tracker = 1

            if settings_dict['saveData'] == True:
                tet_file = session.session_metadata.file_paths['tet']

                # Session stamp
                signature = tet_file.split("/")[-1][:-2]

                if settings['naming_type'] == 'LEC':
                    group, name = extract_name_lec(signature)
                    formats = LEC_naming_format[group][name]['object']
                elif settings['naming_type'] == 'MEC':
                        name = extract_name_mec(signature)
                        formats = MEC_naming_format
                # elif settings['naming_type'] == 'LC':
                #     name = extract_name_lc(fname)
                #     formats = LC_naming_format

                for format in list(formats.keys()):
                    checked = _check_single_format(signature, format, formats[format])
                    if checked is not None:
                        break
                    else:
                        continue
                
                stim, depth, name, date = checked

                # Create save_folder
                save_path = os.path.join(root_path, 'session_' + signature) 

                # if not os.path.isdir(save_path):
                #     os.mkdir(save_path)

                pt = os.path.join(root_path,'shuffled_dist_plots')
                if not os.path.isdir(pt):
                    os.mkdir(pt)

                directory = 'Tetrode_' + tet_file[-1]

                # # Building save directories for tetrodes
                # tetrode_directory_paths = dict()
                # # for directory in tetrode_directories:
                # path = os.path.join(save_path, directory)

                # if not os.path.isdir(path):
                #     os.mkdir(path)
                # tetrode_directory_paths[directory] = path

                # # CREATING EXCEL SHEET FOR STATISTIC SCORES
                # wb = animal_workbooks[animal_id]

                if settings_dict['saveMethod'] == 'one_per_session':
                    wb = animal_workbooks[animal_id][session_key]
                    # sum_sheet = wb['Sheet']
                    # sum_sheet.title = 'Summary'
                    current_statistics_sheet = wb.create_sheet(title=directory)
                    # Copy headers into excel file
                    #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                    current_statistics_sheet['A' + str(1)] = 'Session'
                    current_statistics_sheet['B' + str(1)] = 'Tetrode'
                    current_statistics_sheet['C' + str(1)] = 'Cell ID'
                    for header in headers:
                        #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                        current_statistics_sheet[headers_dict[header] + str(1)] = header
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    if directory not in wb.sheetnames:
                        current_statistics_sheet = wb.create_sheet(title=directory)
                        # Copy headers into excel file
                        #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                        current_statistics_sheet['A' + str(1)] = 'Session'
                        current_statistics_sheet['B' + str(1)] = 'Tetrode'
                        current_statistics_sheet['C' + str(1)] = 'Cell ID'
                        for header in headers:
                            #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                            current_statistics_sheet[headers_dict[header] + str(1)] = header
                    else:
                        current_statistics_sheet = wb[str(directory)]
                elif settings_dict['saveMethod'] == 'one_for_parent':
                    current_statistics_sheet = wb['Summary']

            for cell in session.get_cell_data()['cell_ensemble'].cells:

                if settings_dict['saveMethod'] == 'one_for_parent':
                    excel_cell_index = per_animal_tracker
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    excel_cell_index = per_animal_tetrode_tracker
                elif settings_dict['saveMethod'] == 'one_per_session':
                    excel_cell_index = per_session_tracker

                current_statistics_sheet['C' + str(excel_cell_index+1)] = cell.cluster.cluster_label
                current_statistics_sheet['B' + str(excel_cell_index+1)] = tet_file[-1]
                # animal_tet_count[animal_id]
                current_statistics_sheet['A' + str(excel_cell_index+1)] = signature

                if all_cells == True or ((all_cells==False) & (cell.cluster.cluster_label >= settings_dict['start_cell'] & cell.cluster.cluster_label <= settings_dict['end_cell'])):
                    print('animal: ' + str(animalID) + ', session: ' + str(k) + ', cell_cluster_label: ' + str(cell.cluster.cluster_label))

                    spatial_spike_train = session.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj, 'speed_bounds': (settings_dict['speed_lowerbound'], settings_dict['speed_upperbound'])})

                    pos_x, pos_y, pos_t, arena_size = spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t, spatial_spike_train.arena_size
                    v = _speed2D(spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t)

                    rate_obj = spatial_spike_train.get_map('rate')
                    rate_map, rate_map_raw = rate_obj.get_rate_map(new_size=settings_dict['ratemap_dims'][0])
                    occ_obj = spatial_spike_train.get_map('occupancy')
                    occ_map, _, _ = occ_obj.get_occupancy_map(new_size=settings_dict['ratemap_dims'][0])
                    # if settings['normalizeRate']:
                    #     rate_map, _ = rate_obj.get_rate_map()
                    # else:
                    #     _, rate_map = rate_obj.get_rate_map()

                    # spikex, spikey, spiket = spatial_spike_train.get_spike_positions()

                    # print('Map Stats')
                    ratemap_stats_dict  = rate_map_stats(spatial_spike_train)
                    coherence = rate_map_coherence(spatial_spike_train)
                    spatial_information_content = ratemap_stats_dict['spatial_information_content']
                    current_statistics_sheet[headers_dict['information'] + str(excel_cell_index+1)] = spatial_information_content
                    current_statistics_sheet[headers_dict['sparsity'] + str(excel_cell_index+1)] = ratemap_stats_dict['sparsity']
                    current_statistics_sheet[headers_dict['selectivity'] + str(excel_cell_index+1)] = ratemap_stats_dict['selectivity']
                    current_statistics_sheet[headers_dict['coherence'] + str(excel_cell_index+1)] = coherence
                    

                    current_statistics_sheet[headers_dict['date'] + str(excel_cell_index+1)] = date 
                    current_statistics_sheet[headers_dict['name'] + str(excel_cell_index+1)] = name
                    current_statistics_sheet[headers_dict['depth'] + str(excel_cell_index+1)] = depth  
                    current_statistics_sheet[headers_dict['stim'] + str(excel_cell_index+1)] = stim

                    shuffled_rate_maps, shuffled_rate_map_raw = rate_obj.get_rate_map(new_size=rate_map.shape[0], shuffle=True, n_repeats=1000)

                    shuffled_information_content = []
                    shuffled_sparsity = []
                    shuffled_selectivity = []
                    shuffled_coherence = []
                    for i in range(len(shuffled_rate_maps)):
                        shuffled_map = shuffled_rate_maps[i]
                        shuffled_map_raw = shuffled_rate_map_raw[i]
                        ratemap_stats_dict  = rate_map_stats(None, ratemap=shuffled_map, occmap=occ_map, override=True)
                        spatial_information_content = ratemap_stats_dict['spatial_information_content']
                        shuffled_information_content.append(spatial_information_content)
                        shuffled_sparsity.append(ratemap_stats_dict['sparsity'])
                        shuffled_selectivity.append(ratemap_stats_dict['selectivity'])
                        shuffled_coherence.append(rate_map_coherence(shuffled_map_raw, smoothing_factor=3))


                    current_statistics_sheet[headers_dict['shuffled_information_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_information_content)
                    current_statistics_sheet[headers_dict['shuffled_information_std'] + str(excel_cell_index+1)] = np.std(shuffled_information_content)
                    current_statistics_sheet[headers_dict['shuffled_sparsity_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_sparsity)
                    current_statistics_sheet[headers_dict['shuffled_sparsity_std'] + str(excel_cell_index+1)] = np.std(shuffled_sparsity)
                    current_statistics_sheet[headers_dict['shuffled_selectivity_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_selectivity)
                    current_statistics_sheet[headers_dict['shuffled_selectivity_std'] + str(excel_cell_index+1)] = np.std(shuffled_selectivity)
                    current_statistics_sheet[headers_dict['shuffled_coherence_mean'] + str(excel_cell_index+1)] = np.mean(shuffled_coherence)
                    current_statistics_sheet[headers_dict['shuffled_coherence_std'] + str(excel_cell_index+1)] = np.std(shuffled_coherence)

                    p_value_information = (np.sum(shuffled_information_content < spatial_information_content)) / (len(shuffled_information_content))
                    p_value_sparsity = (np.sum(shuffled_sparsity > ratemap_stats_dict['sparsity'])) / (len(shuffled_sparsity))
                    p_value_selectivity = (np.sum(shuffled_selectivity < ratemap_stats_dict['selectivity'])) / (len(shuffled_selectivity))
                    p_value_coherence = (np.sum(shuffled_coherence < coherence)) / (len(shuffled_coherence))


                    dists = [shuffled_information_content, shuffled_sparsity, shuffled_selectivity, shuffled_coherence]
                    quants = [spatial_information_content, ratemap_stats_dict['sparsity'], ratemap_stats_dict['selectivity'], coherence]
                    order = ['Information','Sparsity','Selectivity','Coherence']
                    file_end = str(animal_id) + '_' + str(date) + '_' + str(session_key) + str(tet_file[-1]) + str(cell.cluster.cluster_label) + '.png'

                    plot_shuffled_dist(root_path, dists, quants, order, file_end)


                    current_statistics_sheet[headers_dict['p_value_information'] + str(excel_cell_index+1)] = p_value_information
                    current_statistics_sheet[headers_dict['p_value_sparsity'] + str(excel_cell_index+1)] = p_value_sparsity
                    current_statistics_sheet[headers_dict['p_value_selectivity'] + str(excel_cell_index+1)] = p_value_selectivity
                    current_statistics_sheet[headers_dict['p_value_coherence'] + str(excel_cell_index+1)] = p_value_coherence

                    sptimes = spatial_spike_train.spike_times
                    t_stop = max(sptimes)
                    t_start = min(sptimes)
                    offset_lim = 20/60
                    if offset_lim >= 0.5 * (t_stop - t_start):
                        offset_lim = int(0.5 * (t_stop - t_start))
                        if offset_lim == 0.5 * (t_stop - t_start):
                            offset_lim -= 1
                    
                    current_statistics_sheet[headers_dict['shuffled_offset'] + str(excel_cell_index+1)] = offset_lim

                    # print('Check Disk')
                    # if 'disk_arena' in tasks and tasks['disk_arena'] == False:
                    fp = session.session_metadata.file_paths['cut']
                    possible_names = ['round', 'cylinder', 'circle']
                    isDisk = False
                    for nm in possible_names:
                        if nm in fp.lower():
                            # isDisk = True
                            tasks['disk_arena'] = True

                        # Auto-resize columns to width of header text
                        #current_statistics_sheet.autofit(axis="columns")
                    ColumnDimension(current_statistics_sheet, bestFit=True)
                    print("Cell " + str(excel_cell_index) + " is complete")
                    per_session_tracker += 1
                    per_animal_tracker += 1
                    per_animal_tetrode_tracker += 1
                #     # -------------------------------------------------------- #

                c += 1

            k += 1

            if settings_dict['saveMethod'] == 'one_per_session':
                if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                    list(map(lambda x: _save_wb(animal_workbooks[animal_id][x], save_path), animal_workbooks[animal_id]))

        if settings_dict['saveMethod'] == 'one_per_animal_tetrode':
            if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                _save_wb(wb, root_path, animal_id=animal_id)

    if settings_dict['saveMethod'] == 'one_for_parent':
        _save_wb(wb, root_path, sum_sheet_count=sum_sheet_count)


def plot_shuffled_dist(root_path, dists, quants, order, file_end):
    fig = plt.figure(figsize=(12,6))

    c = 0 
    for i in [1,3,5,7]:
        ax = plt.subplot(4,2,i)
        out = ax.hist(dists[c], bins=100, color='grey')
        ax.vlines(quants[c],0,np.max(out[0]), color='r')
        ax.set_xlabel(order[c])
        ax.set_ylabel('Count')
        ax.set_title('Non-log')
        

        ax_log = plt.subplot(4,2,i+1)
        try:
            out = ax_log.hist(np.log(dists[c]), bins=100, color='k')
            ax_log.vlines(np.log(quants[c]),0,np.max(out[0]), color='r')
        except:
            pass
        ax_log.set_xlabel(order[c])
        ax_log.set_ylabel('Count')
        ax_log.set_title('Post log')

        c += 1

    fig.tight_layout()

    pth = root_path + '/shuffled_dist_plots/' + str(file_end)

    fig.savefig(pth)

    plt.close()



def _save_wb(wb, root_path, animal_id=None, sum_sheet_count=None):
    wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
    if animal_id is None:
        if sum_sheet_count is None:
            pth = root_path + '/shuffle_sheet'  + '.xlsx'
        else:
            pth = root_path + '/shuffle_sheet_'  + str(sum_sheet_count) + '.xlsx'
    else:
        pth = root_path + '/shuffle_sheet_' + str(animal_id)  + '.xlsx'
    print(root_path)
    wb.save(pth)
    wb.close()

    xls = pd.read_excel(pth, sheet_name=None)
    df_sum = xls.pop('Summary')
    dfs = [df.sort_values(['Session', 'Tetrode', 'Cell ID']) for df in xls.values()]
    with pd.ExcelWriter(pth, engine='xlsxwriter') as writer:
        df_sum.to_excel(writer, sheet_name='Summary', index=False)
        for sheet, df in zip(xls.keys(), dfs):
            df.to_excel(writer, sheet_name=sheet, index=False)


    
        if len(dfs) > 0:
            df_sum = pd.concat(dfs, axis=0).sort_values(['Session', 'Tetrode', 'Cell ID'])
        else:
            df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell ID'])
        df_sum.to_excel(writer, sheet_name="Summary", index=True)
    # writer.save()
    print('Saved ' + str(pth))

def get_hd_score_for_cluster(hd_hist):
    angles = np.linspace(-179, 180, 360)
    angles_rad = angles*np.pi/180
    dy = np.sin(angles_rad)
    dx = np.cos(angles_rad)

    totx = sum(dx * hd_hist)/sum(hd_hist)
    toty = sum(dy * hd_hist)/sum(hd_hist)
    r = np.sqrt(totx*totx + toty*toty)
    return r

if __name__ == '__main__':

    ######################################################## EDIT BELOW HERE ########################################################

    # MAKE SURE "field_sizes" IS THE LAST ELEMENT IN "csv_header_keys"
    csv_header = {}
    csv_header_keys = ['name','date','depth','stim','information', 'shuffled_information_mean','shuffled_information_std','p_value_information',
                       'selectivity', 'shuffled_selectivity_mean', 'shuffled_selectivity_std', 'p_value_selectivity',
                       'sparsity', 'shuffled_sparsity_mean', 'shuffled_sparsity_std', 'p_value_sparsity',
                       'coherence', 'shuffled_coherence_mean', 'shuffled_coherence_std', 'p_value_coherence',
                       'shuffled_offset']
    for key in csv_header_keys:
        csv_header[key] = True

    tasks = {}
    task_keys = ['information']
    for key in task_keys:
        tasks[key] = True

    plotTasks = {}
    plot_task_keys = ['Spikes_Over_Position_Map', 'Tuning_Curve_Plots', 'Firing_Rate_vs_Speed_Plots', 'Firing_Rate_vs_Time_Plots','autocorr_map', 'binary_map','rate_map', 'occupancy_map']
    for key in plot_task_keys:
        plotTasks[key] = True

    animal = {'animal_id': '001', 'species': 'mouse', 'sex': 'F', 'age': 1, 'weight': 1, 'genotype': 'type', 'animal_notes': 'notes'}
    devices = {'axona_led_tracker': True, 'implant': True}
    implant = {'implant_id': '001', 'implant_type': 'tetrode', 'implant_geometry': 'square', 'wire_length': 25, 'wire_length_units': 'um', 'implant_units': 'uV'}

    session_settings = {'channel_count': 4, 'animal': animal, 'devices': devices, 'implant': implant}

    """ FOR YOU TO EDIT """
    settings = {'ppm': 485, 'session':  session_settings, 'smoothing_factor': 3, 'useMatchedCut': True}
    """ FOR YOU TO EDIT """

    tasks['disk_arena'] = True # -->
    settings['tasks'] = tasks # --> change tasks array to change tasks are run
    settings['plotTasks'] = plotTasks # --> change plot tasks array to change asks taht are plotted
    settings['header'] = csv_header # --> change csv_header header to change tasks that are saved to csv

    """ FOR YOU TO EDIT """
    settings['naming_type'] = 'LEC'
    settings['speed_lowerbound'] = 0
    settings['speed_upperbound'] = 99
    settings['end_cell'] = None
    settings['start_cell'] = None
    settings['saveData'] = True
    settings['ratemap_dims'] = (32, 32)
    settings['saveMethod'] = 'one_for_parent'
    # possible saves are:
    # 1 csv per session (all tetrode and indiv): 'one_per_session' --> 5 sheets (summary of all 4 tetrodes, tet1, tet2, tet3, tet4)
    # 1 csv per animal per tetrode (all sessions): 'one_per_animal_tetrode' --> 4 sheets one per tet 
    # 1 csv per animal (all tetrodes & sessions): 'one_for_parent' --> 1 sheet
    """ FOR YOU TO EDIT """


    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    data_dir = filedialog.askdirectory(parent=root,title='Please select a data directory.')

    ########################################################################################################################

    """ OPTION 1 """
    """ RUNS EVERYTHING UNDER PARENT FOLDER (all subfolders loaded first) """
    # study = make_study(data_dir,settings_dict=settings)
    # study.make_animals()
    # batch_map(study, settings, data_dir)

    """ OPTION 2 """
    """ RUNS EACH SUBFOLDER ONE AT A TIME """
    subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    count = 1
    for subdir in subdirs:
        try:
            study = make_study(subdir,settings_dict=settings)
            study.make_animals()
            batch_map(study, settings, subdir, sum_sheet_count=count)
            count += 1
        except Exception:
            print(traceback.format_exc())
            print('DID NOT WORK FOR DIRECTORY ' + str(subdir))

    # """ OPTION 3 """
    # """ RUNS EACH SUBFOLDER ONE AT A TIME """
    # subdirs = np.sort([ f.path for f in os.scandir(data_dir) if f.is_dir() ])
    # count = 1
    # for subdir in subdirs:
    #     subdir = str(subdir)
    #     subdirs2 = np.sort([ f.path for f in os.scandir(subdir) if f.is_dir() ])
    #     for subdir2 in subdirs2:
    #         try:
    #             study = make_study(subdir2,settings_dict=settings)
    #             study.make_animals()
    #             batch_map(study, settings, subdir2, sum_sheet_count=count)
    #             count += 1
    #         except Exception:
    #             print(traceback.format_exc())
    #             print('DID NOT WORK FOR DIRECTORY ' + str(subdir2))