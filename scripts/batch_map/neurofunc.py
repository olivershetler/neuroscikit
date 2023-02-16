import os
import sys

PROJECT_PATH = os.getcwd()
sys.path.append(PROJECT_PATH)

from _prototypes.unit_matcher.waveform import time_index, troughs
import warnings
from library.study_space import Session, Study, Animal
from library.workspace import Workspace
import tkinter as tk
from tkinter import filedialog
import time
from library.maps import autocorrelation, binary_map, spatial_tuning_curve, map_blobs
from library.hafting_spatial_maps import HaftingOccupancyMap, HaftingRateMap, HaftingSpikeMap, SpatialSpikeTrain2D
from library.scores import hd_score, grid_score, border_score
from library.scores import rate_map_stats, rate_map_coherence, speed_score
from openpyxl.utils.cell import get_column_letter, column_index_from_string
from library.maps.map_utils import disk_mask
from PIL import Image
import numpy as np
from matplotlib import cm
import matplotlib.pyplot as plt
#import random
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension
from matplotlib import cm
# from opexebo.analysis import rate_map_stats, speed_score, rate_map_coherence
# Set matplotlib backend to Agg to prevent figures from popping up.
from matplotlib import use
from library.maps.firing_rate_vs_time import firing_rate_vs_time
from library.maps.filter_pos_by_speed import filter_pos_by_speed
from library.map_utils import _speed2D
from x_io.rw.axona.batch_read import make_study
import pandas as pd
from library.cluster import create_features, L_ratio, isolation_distance
from library.spike import histogram_ISI, find_burst


# class BatchMaps(Workspace):
#     def __init__(self, study: StudyWorkspace):
#         pass

# def add_to_sheet(to_add, headers_dict, header, wb, method, sheet, cell_index):
#     current_statistics_sheet[headers_dict['border_score_top'] + str(excel_cell_index+1)] = b_score[0]

#     if method == 'one_per_session':
#         sheet[headers_dict[header] + str(cell_index+1)] = to_add
#     elif method == 'one_per_animal_tetrode':
#         sheet[headers_dict[header] + str(cell_index+1)] = to_add

""" SETTINGS AT THE BOTTOM OF FILE """


def batch_map(study: Study, settings_dict: dict, saveDir=None):
    """
    Computes rate maps across all animals, sessions, cells in a study.

    Use tasks dictionary as true/false flag with variable to compute
    e.g. {'rate_map': True, 'binary_map': False}
    """

    tasks = settings_dict['tasks']
    plotTasks = settings_dict['plotTasks']
    csv_header = settings_dict['header']

    # Root path creation
    save_dir = saveDir
    root_path = os.path.join(save_dir, 'All_PRISM_Sessions_iter_')
    run_number = 1
    while os.path.isdir(root_path+str(run_number)):
        run_number+=1

    root_path += str(run_number)
    os.mkdir(root_path)

    # Kernel size
    kernlen = int(settings_dict['smoothing_factor']*8)
    # Standard deviation size
    std = int(0.2*kernlen)

    # Set flag if no arguments were provided for later
    all_cells = False
    if (settings_dict['start_cell'] == settings_dict['end_cell'] == None):
        all_cells = True

    # Grabs headers whose value is true
    headers = [k for k, v in csv_header.items() if v]
    if 'border_score' in headers:
        headers.remove('border_score')
        headers.insert(5, 'border_score_top')
        headers.insert(6, 'border_score_bottom')
        headers.insert(7, 'border_score_left')
        headers.insert(8, 'border_score_right')

    headers_dict = dict()

    for i, header in enumerate(headers):
        headers_dict[header] = get_column_letter(i+4)

    file_name = os.path.join(root_path, "PRISM_parameters.txt")
    with open(file_name, 'w') as f:
        f.write("ppm is: ")
        f.write(str(settings_dict['ppm']) + "\n")
        f.write("Smoothing is: ")
        f.write(str(settings_dict['smoothing_factor']) + "\n")
        f.write("speed bounds are: ")
        f.write(str(settings_dict['speed_lowerbound']) + "," + str(settings_dict['speed_upperbound']) + "\n")
        f.write("save method is: ")
        f.write(str(settings_dict['saveMethod']) + "\n")
        f.close()

    if study.animals is None:
        study.make_animals()
        print('Animals made, batching map')

    animals = study.animals

    analysis_directories = [k for k, v in plotTasks.items() if v]

    animal_tet_count = {}
    animal_max_tet_count = {}
    animal_workbooks = {}

    per_animal_tracker = 1
    per_animal_tetrode_tracker = 1

    sorted_animal_ids = np.unique(np.sort(study.animal_ids))

    for animalID in sorted_animal_ids:
        animal = study.get_animal_by_id(animalID)

        animal_id = animal.animal_id.split('_tet')[0]
        if animal_id not in animal_tet_count:
            animal_tet_count[animal_id] = 1
            wb = xl.Workbook()
            animal_workbooks[animal_id] = wb
            animal_max_tet_count[animal_id] = len(sorted_animal_ids[list(map(lambda x: True if animal_id in x else False, sorted_animal_ids))])
            sum_sheet = wb['Sheet']
            sum_sheet.title = 'Summary'
            if settings_dict['saveMethod'] == 'one_per_session':
                # animal_tet_count[animal_id] = np.ones(len(animal.sessions))
                animal_workbooks[animal_id] = {}
                for ses_key in animal.sessions:
                    wb = xl.Workbook()
                    sum_sheet = wb['Sheet']
                    sum_sheet.title = 'Summary'
                    animal_workbooks[animal_id][ses_key] = wb
            elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                # wb.remove_sheet('Summary') 
                pass
            elif settings_dict['saveMethod'] == 'one_per_animal':
                # Copy headers into excel file
                #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                sum_sheet['A' + str(1)] = 'Session'
                sum_sheet['B' + str(1)] = 'Tetrode'
                sum_sheet['C' + str(1)] = 'Cell'
                for header in headers:
                    #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                    sum_sheet[headers_dict[header] + str(1)] = header
        else:
            animal_tet_count[animal_id] += 1
            wb = animal_workbooks[animal_id]

        # cells, waveforms = sort_cell_spike_times(animal)

        # sort_cell_spike_times(animal)

        # cluster = session.get_spike_data['spike_cluster']

        # sort_spikes_by_cell(cluster)

        k = 1

        # col_count = 0 

        for session_key in animal.sessions:
            session = animal.sessions[session_key]

            c = 0

            pos_obj = session.get_position_data()['position']

            # excel_cell_index = 1
            per_session_tracker = 1

            if settings_dict['saveData'] == True:
                tet_file = session.session_metadata.file_paths['tet']

                # Session stamp
                signature = tet_file.split("/")[-1][:-2]

                # Create save_folder
                save_path = os.path.join(root_path, 'PRISM_Session_' + signature) + '_' + str(c)

                if not os.path.isdir(save_path):
                    os.mkdir(save_path)
                # animal_tet_count[animal_id] += 1

                directory = 'Tetrode_' + tet_file[-1]
                        
                # Creating file directories for plots
                root_directory_paths = dict()
                root_directory_paths['Occupancy_Map'] = os.path.join(save_path, 'Occupancy_Map')

                # Creating occupancy map directory
                if not os.path.isdir(root_directory_paths['Occupancy_Map']):
                    os.mkdir(root_directory_paths['Occupancy_Map'])

                # Building save directories for tetrodes
                tetrode_directory_paths = dict()
                # for directory in tetrode_directories:
                path = os.path.join(save_path, directory)

                if not os.path.isdir(path):
                    os.mkdir(path)
                tetrode_directory_paths[directory] = path

                # # CREATING EXCEL SHEET FOR STATISTIC SCORES
                # wb = animal_workbooks[animal_id]

                if settings_dict['saveMethod'] == 'one_per_session':
                    wb = animal_workbooks[animal_id][session_key]
                    # sum_sheet = wb['Sheet']
                    # sum_sheet.title = 'Summary'
                    current_statistics_sheet = wb.create_sheet(title=directory)
                    # Copy headers into excel file
                    #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                    current_statistics_sheet['A' + str(1)] = 'Session'
                    current_statistics_sheet['B' + str(1)] = 'Tetrode'
                    current_statistics_sheet['C' + str(1)] = 'Cell'
                    for header in headers:
                        #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                        current_statistics_sheet[headers_dict[header] + str(1)] = header
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    if directory not in wb.sheetnames:
                        current_statistics_sheet = wb.create_sheet(title=directory)
                        # Copy headers into excel file
                        #current_statistics_sheet.range('A' + str(1)).value = 'Cell'
                        current_statistics_sheet['A' + str(1)] = 'Session'
                        current_statistics_sheet['B' + str(1)] = 'Tetrode'
                        current_statistics_sheet['C' + str(1)] = 'Cell'
                        for header in headers:
                            #current_statistics_sheet.range(get_column_letter(i+2) + str(1)).value = value
                            current_statistics_sheet[headers_dict[header] + str(1)] = header
                    else:
                        current_statistics_sheet = wb[str(directory)]
                elif settings_dict['saveMethod'] == 'one_per_animal':
                    current_statistics_sheet = wb['Summary']
            
            if tasks['spike_analysis']:
                cluster_batch = session.get_spike_data()['spike_cluster']
                create_features(cluster_batch)
                L_ratio(cluster_batch)
                isolation_distance(cluster_batch)
                

            for cell in session.get_cell_data()['cell_ensemble'].cells:

                if settings_dict['saveMethod'] == 'one_per_animal':
                    excel_cell_index = per_animal_tracker
                elif settings_dict['saveMethod'] == 'one_per_animal_tetrode':
                    excel_cell_index = per_animal_tetrode_tracker
                elif settings_dict['saveMethod'] == 'one_per_session':
                    excel_cell_index = per_session_tracker

                current_statistics_sheet['C' + str(excel_cell_index+1)] = cell.cluster.cluster_label
                current_statistics_sheet['B' + str(excel_cell_index+1)] = tet_file[-1]
                # animal_tet_count[animal_id]
                current_statistics_sheet['A' + str(excel_cell_index+1)] = signature

                if all_cells == True or ((all_cells==False) & (cell.cluster.cluster_label >= settings_dict['start_cell'] & cell.cluster.cluster_label <= settings_dict['end_cell'])):
                    print('animal: ' + str(animalID) + ', session: ' + str(k) + ', cell_cluster_label: ' + str(cell.cluster.cluster_label))
                    # print(cell.event_times[:10])
                    # print('SpatialSPikeTrain Class')
                    spatial_spike_train = session.make_class(SpatialSpikeTrain2D, {'cell': cell, 'position': pos_obj, 'speed_bounds': (settings_dict['speed_lowerbound'], settings_dict['speed_upperbound'])})

                    pos_x, pos_y, pos_t, arena_size = spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t, spatial_spike_train.arena_size
                    v = _speed2D(spatial_spike_train.x, spatial_spike_train.y, spatial_spike_train.t)
                    # stop()

                    # # print('HafftingOccupancyMap')
                    # occ_obj = HaftingOccupancyMap(spatial_spike_train)
                    # occ_map, _, _ = occ_obj.get_occupancy_map()

                    # # print('HaftingSpikeMap')
                    # spike_obj = HaftingSpikeMap(spatial_spike_train)
                    # spike_map = spike_obj.get_spike_map()

                    # print('HaftingRateMap')
                    # rate_obj = HaftingRateMap(spatial_spike_train)
                    rate_obj = spatial_spike_train.get_map('rate')
                    rate_map, rate_map_raw = rate_obj.get_rate_map()
                    # if settings['normalizeRate']:
                    #     rate_map, _ = rate_obj.get_rate_map()
                    # else:
                    #     _, rate_map = rate_obj.get_rate_map()

                    # spikex, spikey, spiket = spatial_spike_train.get_spike_positions()



                    # print('Map Stats')
                    ratemap_stats_dict  = rate_map_stats(spatial_spike_train)

                    # UNDO COMMENT
                    autocorr_map = autocorrelation(spatial_spike_train)

                    occ_obj = spatial_spike_train.get_map('occupancy')
                    occ_map, _, _ = occ_obj.get_occupancy_map()
                    spike_obj = spatial_spike_train.get_map('spike')
                    spike_map, _ = spike_obj.get_spike_map()

                    spikex, spikey, spiket = spike_obj.spike_x, spike_obj.spike_y, spike_obj.new_spike_times

                    cell_stats = {}
                    cell_stats['rate_map_smooth'] = rate_map
                    cell_stats['occupancy_map'] = occ_map
                    cell_stats['rate_map_raw'] = rate_map_raw
                    cell_stats['autocorrelation_map'] = autocorr_map
                    cell_stats['spatial_spike_train'] = spatial_spike_train

                    if tasks['spike_analysis']:
                        bursting, avg_spikes_per_burst = find_burst(cell)
                        ISI_dict = histogram_ISI(cell)
                    
                    if tasks['spike_width']:
                        # n_spikes, spike_times, waveforms = session.session_data.data['spike_cluster'].get_single_spike_cluster_instance(unit)
                        n_spikes = len(cell.event_times)
                        waveforms = cell.signal
                        wf_avg = np.array(waveforms).mean(axis=1)
                        max_vals = list(map(lambda x: max(x), [wf_avg[0], wf_avg[1], wf_avg[2], wf_avg[3]]))
                        principal_channel_index = np.argmax(max_vals)
                        principal_waveform = wf_avg[principal_channel_index]
                        peak_index = np.argmax(principal_waveform)
                        trough_list = list(filter(lambda x: x > peak_index, troughs(principal_waveform)))
                        sample_rate = session.session_data.data['spike_cluster'].waveform_sample_rate
                        if len(trough_list) > 0:
                            trough_index = trough_list[0]
                            spike_width = (trough_index - peak_index) / sample_rate
                        else:
                            trough_index = len(principal_waveform) - 1
                            spike_width = int((trough_index - peak_index)/2) / sample_rate
                        if spike_width < 0:
                            warnings.warn(f'Negative spike width for unit {cell.cluster.cluster_label} in session {signature}.\n\nThe mean waveform is:\n{principal_waveform}\n\nThe peak index is {peak_index} and the trough index is {trough_index}. The spike width is {spike_width}.\n\nThe sample_rate is {sample_rate}.')
                        duration = session.session_data.data['spike_cluster'].duration
                        firing_rate = n_spikes/duration
                        cell_stats['firing_rate'] = firing_rate
                        cell_stats['spike_width'] = spike_width

                    # print('Check Disk')
                    if 'disk_arena' in tasks and tasks['disk_arena'] == False:
                        fp = session.session_metadata.file_paths['cut']
                        possible_names = ['round', 'Round', 'ROUND', 'Cylinder', 'cylinder', 'CYLINDER', 'circle', 'CIRCLE', 'Circle']
                        for name in possible_names:
                            if name in fp:
                                tasks['disk_arena'] = True

                    # print('Binary')
                    if tasks['binary_map']:
                        binmap = binary_map(spatial_spike_train)
                        if tasks['disk_arena']:
                            binmap = disk_mask(binmap)
                        # binmap_im = Image.fromarray(np.uint8(binmap*255))
                        cell_stats['binary_map'] = binmap
                        # cell_stats['binary_map_im'] = binmap_im

                    # print('Autocorr Img')
                    if tasks['autocorrelation_map']:
                        cell_stats['autocorr_map'] = autocorr_map
                        if tasks['disk_arena']:
                            autocorr_map = disk_mask(autocorr_map)
                        # autocorr_map_im = Image.fromarray(np.uint8(cm.jet(autocorr_map)*255))
                        # cell_stats['autocorr_map_im'] = autocorr_map_im

                    if tasks['sparsity']:
                        cell_stats['sparsity'] = ratemap_stats_dict['sparsity']

                    if tasks['selectivity']:
                        cell_stats['selectivity'] = ratemap_stats_dict['selectivity']

                    if tasks['information']:
                        cell_stats['spatial_information_content'] = ratemap_stats_dict['spatial_information_content']

                    # print('Coherence')
                    if tasks['coherence']:
                        coherence = rate_map_coherence(spatial_spike_train)
                        cell_stats['coherence'] = coherence

                    # print('Speed Score')
                    if tasks['speed_score']:
                        s_score = speed_score(spatial_spike_train)
                        cell_stats['speed_score'] = s_score

                    # print('Spatial Tuning Curve')
                    if tasks['hd_score'] or tasks['tuning_curve']:
                        tuned_data, spike_angles, angular_occupancy, bin_array = spatial_tuning_curve(spatial_spike_train)
                        cell_stats['tuned_data'] = tuned_data
                        cell_stats['tuned_data_angles'] = spike_angles
                        cell_stats['angular_occupancy'] = angular_occupancy
                        cell_stats['angular_occupancy_bins'] = bin_array

                    # print('HD score')
                    if tasks['hd_score']:
                        hd_hist = hd_score(spatial_spike_train)
                        cell_stats['hd_hist'] = hd_hist

                    # print('Grid score')
                    if tasks['grid_score']:
                        true_grid_score = grid_score(spatial_spike_train)
                        cell_stats['grid_score'] = true_grid_score

                    # print('Border score')
                    if tasks['border_score'] and not tasks['disk_arena']:
                        b_score = border_score(spatial_spike_train)
                        cell_stats['b_score_top'] = b_score[0]
                        cell_stats['b_score_bottom'] = b_score[1]
                        cell_stats['b_score_left'] = b_score[2]
                        cell_stats['b_score_right'] = b_score[3]
                    elif tasks['border_score'] and tasks['disk_arena']:
                        print('Cannot compute border score on disk arena')

                    # print('Field sizes')
                    if tasks['field_sizes']:
                        image, n_labels, labels, centroids, field_sizes = map_blobs(spatial_spike_train)
                        cell_stats['field_size_data'] = {'image': image, 'n_labels': n_labels, 'labels': labels, 'centroids': centroids, 'field_sizes': field_sizes}

                    cell.stats_dict['cell_stats'] = cell_stats

                    if settings_dict['saveData'] == True:

                        if plotTasks['rate_map']:
                            colored_ratemap = Image.fromarray(np.uint8(cm.jet(rate_map)*255))
                            colored_ratemap.save('ratemap_cell_' + str(c) + '_session_' + str(k) + '.png')
                        
                        if tasks['disk_arena'] and plotTasks['occupancy_map']:
                            cell_stats['occupancy_map'] = disk_mask(cell_stats['occupancy_map'])
                            colored_occupancy_map = Image.fromarray(np.uint8(cm.jet(cell_stats['occupancy_map'])*255))
                            colored_occupancy_map.save(root_directory_paths['Occupancy_Map'] + '/pospdf_cell' + '.png')

                        # Binary ratemap
                        if plotTasks['binary_map']:
                            im = Image.fromarray(np.uint8(binmap*255))
                            im.save(tetrode_directory_paths[directory] + '/Binary_Map_Cell_' + str(i) + '.png')

                        # Sparsity, Selectivity and Shannon
                        if tasks['sparsity']:
                            #current_statistics_sheet.range(headers_dict['Sparsity'] + str(excel_cell_index+1)).value = ratemap_stats_dict['sparsity']
                            current_statistics_sheet[headers_dict['sparsity'] + str(excel_cell_index+1)] = ratemap_stats_dict['sparsity']


                        if tasks['selectivity']:
                            #current_statistics_sheet.range(headers_dict['Selectivity'] + str(excel_cell_index+1)).value = ratemap_stats_dict['selectivity']
                            current_statistics_sheet[headers_dict['selectivity'] + str(excel_cell_index+1)] = ratemap_stats_dict['selectivity']


                        if tasks['information']:
                            #current_statistics_sheet.range(headers_dict['Information'] + str(excel_cell_index+1)).value = ratemap_stats_dict['spatial_information_content']
                            current_statistics_sheet[headers_dict['information'] + str(excel_cell_index+1)] = ratemap_stats_dict['spatial_information_content']

                        if tasks['spike_width']:
                            current_statistics_sheet[headers_dict['spike_width'] + str(excel_cell_index+1)] = spike_width
                            current_statistics_sheet[headers_dict['firing_rate'] + str(excel_cell_index+1)] = firing_rate

                        if tasks['spike_analysis']:
                            current_statistics_sheet[headers_dict['bursting'] + str(excel_cell_index+1)] = bursting
                            current_statistics_sheet[headers_dict['Avg. Spikes/Burst'] + str(excel_cell_index+1)] = avg_spikes_per_burst
                            current_statistics_sheet[headers_dict['ISI_min'] + str(excel_cell_index+1)] = ISI_dict['min']
                            current_statistics_sheet[headers_dict['ISI_max'] + str(excel_cell_index+1)] = ISI_dict['max']
                            current_statistics_sheet[headers_dict['ISI_median'] + str(excel_cell_index+1)] = ISI_dict['median']
                            current_statistics_sheet[headers_dict['ISI_mean'] + str(excel_cell_index+1)] = ISI_dict['mean']
                            current_statistics_sheet[headers_dict['ISI_std'] + str(excel_cell_index+1)] = ISI_dict['std']
                            current_statistics_sheet[headers_dict['ISI_cv'] + str(excel_cell_index+1)] = ISI_dict['cv']
                            

                        # autocorrelation map
                        if plotTasks['autocorr_map']:
                            im = Image.fromarray(np.uint8(cm.jet(autocorr_map)*255))
                            im.save(tetrode_directory_paths[directory] + '/autocorr_cell_' + str(i) + '.png')

                        if tasks['border_score'] and not tasks['disk_arena']:
                            current_statistics_sheet[headers_dict['border_score_top'] + str(excel_cell_index+1)] = b_score[0]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Bottom'] + str(excel_cell_index+1)).value = b_score[1]
                            current_statistics_sheet[headers_dict['border_score_bottom'] + str(excel_cell_index+1)] = b_score[1]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Left'] + str(excel_cell_index+1)).value = b_score[2]
                            current_statistics_sheet[headers_dict['border_score_left'] + str(excel_cell_index+1)] = b_score[2]
                            #current_statistics_sheet.range(headers_dict['Border_Score_Right'] + str(excel_cell_index+1)).value = b_score[3]
                            current_statistics_sheet[headers_dict['border_score_right'] + str(excel_cell_index+1)] = b_score[3]

                        # Coherence
                        if tasks['coherence']:
                            current_statistics_sheet[headers_dict['coherence'] + str(excel_cell_index+1)] = coherence

                        # Speed score
                        if tasks['speed_score']:
                            current_statistics_sheet[headers_dict['speed_score'] + str(excel_cell_index+1)] = s_score[0]['2016']

                        # Head direction scores
                        if tasks['hd_score']:
                            current_statistics_sheet[headers_dict['hd_score'] + str(excel_cell_index+1)] = get_hd_score_for_cluster(hd_hist)

                        # Grid score
                        if tasks['grid_score']:
                            current_statistics_sheet[headers_dict['grid_score'] + str(excel_cell_index+1)] = true_grid_score
                            shuffle_check = True

                        # Field sizes
                        if tasks['field_sizes']:
                            # image, n_labels, labels, centroids, field_sizes = get_map_blobs(rate_map_smooth)
                            # prev = None
                            # prevZ = False
                            # extension_list = ['AA','AB','AC','AD','AE']
                            first_fields_index = column_index_from_string(headers_dict['field_sizes']) + 1

                            for j, field_size in enumerate(cell_stats['field_size_data']['field_sizes']):
                                # if not prevZ:
                                #     if prev == 'Z':
                                #         new_j = 0 
                                #         prevZ = True
                                #         column_letter = extension_list[new_j]
                                #     else:
                                #         column_letter = chr( ord( headers_dict['field_sizes']) + j+1 )
                                # else:
                                col_check_letter = get_column_letter(first_fields_index + j)
                                # print(current_statistics_sheet[col_check_letter + "1"], 'Field_' + str(j+1))
                                if current_statistics_sheet[col_check_letter + "1"] != 'Field_' + str(j+1):
                                    # column_letter = get_column_letter(current_statistics_sheet.max_column + 1)
                                    current_statistics_sheet[col_check_letter + "1"] = 'Field_' + str(j+1)

                                current_statistics_sheet[col_check_letter + str(excel_cell_index+1)] = field_size
                                # print(column_letter, j, headers_dict['field_sizes'])                    
                                # prev = column_letter

                        # Firing rate vs. time, for a window of 400 millieconds
                        if plotTasks['Firing_Rate_vs_Time_Plots']:
                            firing_rate, firing_time = firing_rate_vs_time(spatial_spike_train.spike_times, pos_t, 400)
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.title('Firing Rate vs. Time')
                            plt.xlabel('Time (seconds)')
                            plt.ylabel('Firing Rate (Hertz)')
                            print
                            plt.plot(pos_t, firing_rate, linewidth=0.25)
                            plt.savefig(tetrode_directory_paths[directory] + '/FRvT_cell_' + str(i) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Firing rate vs. speed
                        if plotTasks['Firing_Rate_vs_Speed_Plots']:
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.title('Firing Rate vs. Speed')
                            plt.xlabel('Speed (cm/s)')
                            plt.ylabel('Firing Rate (Hertz)')
                            plt.scatter(v, firing_rate, s = 0.5)
                            plt.savefig(tetrode_directory_paths[directory] + '/FRvS_cell_' + str(i) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Plotting tuning curves (polar plots for directional firing) per cell
                        if plotTasks['Tuning_Curve_Plots']:
                            tuned_data = cell_stats['tuned_data'] 
                            spike_angles = cell_stats['tuned_data_angles']
                            angular_occupancy = cell_stats['angular_occupancy']
                            bin_array = cell_stats['angular_occupancy_bins']
                            last_and_first_averaged = (tuned_data[0]+tuned_data[-1]) / 2
                            tuned_data[0] = last_and_first_averaged
                            tuned_data[-1] = last_and_first_averaged

                            fig, ax = plt.subplots(subplot_kw={'projection': 'polar'})
                            ax.plot(bin_array, tuned_data, linewidth=6)
                            ax.set_xticks(np.arange(0,2.0*np.pi,np.pi/2.0))
                            ax.set_yticks(np.linspace(  0, max(tuned_data), 2)[-1]  )
                            # fig = plt.figure()
                            # figure = plt.gcf()
                            # figure.set_size_inches(4, 4)
                            plt.title("Polar plot")
                            # plt.polar(bin_array, tuned_data, linewidth=3)

                            plt.box(on=None)
                            plt.savefig(tetrode_directory_paths[directory] + '/tuning_curve_cell_' + str(i) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Spikes over position map
                        if plotTasks['Spikes_Over_Position_Map']:
                            fig = plt.figure()
                            figure = plt.gcf()
                            figure.set_size_inches(4, 4)
                            plt.plot(pos_x,pos_y, linewidth=0.2)
                            plt.scatter(spikex,spikey, c='r', s=5, zorder=3)
                            plt.title("Spikes over position")
                            plt.xlabel("x coordinates")
                            plt.ylabel("y coordinates")
                            plt.gca().set_aspect('equal', adjustable='box')
                            plt.savefig(tetrode_directory_paths[directory] + '/spikes_over_position_cell_' + str(i) + '.png', dpi=300, bbox_inches = 'tight')
                            plt.close(fig)

                        # Auto-resize columns to width of header text
                        #current_statistics_sheet.autofit(axis="columns")
                        ColumnDimension(current_statistics_sheet, bestFit=True)

                        print("Cell " + str(excel_cell_index) + " is complete")
                        per_session_tracker += 1
                        per_animal_tracker += 1
                        per_animal_tetrode_tracker += 1
                #     # -------------------------------------------------------- #

                c += 1

            k += 1

            if settings_dict['saveMethod'] == 'one_per_session':
                if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:

                    list(map(lambda x: _save_wb(animal_workbooks[animal_id][x], save_path), animal_workbooks[animal_id]))

        if settings_dict['saveMethod'] == 'one_per_animal_tetrode':
            if animal_tet_count[animal_id] == animal_max_tet_count[animal_id]:
                _save_wb(wb, root_path)

    if settings_dict['saveMethod'] == 'one_per_animal':
        _save_wb(wb, root_path)
        # wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
        # print(root_path)
        # pth = root_path + '/summary_sheet'  + '.xlsx'
        # wb.save(pth)
        # wb.close()

        # xls = pd.ExcelFile(pth)
        # df_sum = pd.read_excel(xls, 'Summary')
        # df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell'])
        # writer = pd.ExcelWriter(pth, engine="xlsxwriter")
        # df_sum.to_excel(writer, sheet_name="Summary", index=False)
        # writer.save()

def _save_wb(wb, root_path):
    wb._sheets = sorted(wb._sheets, key=lambda x: x.title)
    pth = root_path + '/summary_sheet'  + '.xlsx'
    print(root_path)
    wb.save(pth)
    wb.close()

    xls = pd.ExcelFile(pth)
    df_sum = pd.read_excel(xls, 'Summary')
    writer = pd.ExcelWriter(pth, engine="xlsxwriter")
    dfs = []
    for sheet in xls.sheet_names:
        if sheet != 'Summary' and sheet != 'summary':
            df = pd.read_excel(xls, sheet)
            dfs.append(df) 

            df = df.sort_values(['Session', 'Tetrode', 'Cell'])
            df.to_excel(writer, sheet_name=sheet, index=True)
            # writer.save()
    
    if len(dfs) > 0:
        df_sum = pd.concat(dfs, axis=0).sort_values(['Session', 'Tetrode', 'Cell'])
    else:
        df_sum = df_sum.sort_values(['Session', 'Tetrode', 'Cell'])
    df_sum.to_excel(writer, sheet_name="Summary", index=True)
    writer.save()
    print('Saved ' + str(pth))

def get_hd_score_for_cluster(hd_hist):
    angles = np.linspace(-179, 180, 360)
    angles_rad = angles*np.pi/180
    dy = np.sin(angles_rad)
    dx = np.cos(angles_rad)

    totx = sum(dx * hd_hist)/sum(hd_hist)
    toty = sum(dy * hd_hist)/sum(hd_hist)
    r = np.sqrt(totx*totx + toty*toty)
    return r

if __name__ == '__main__':

    ######################################################## EDIT BELOW HERE ########################################################

    # MAKE SURE "field_sizes" IS THE LAST ELEMENT IN "csv_header_keys"
    csv_header = {}
    csv_header_keys = ['spike_width', 'firing_rate', 'Avg. Spikes/Burst', 'bursting', 'ISI_min', 'ISI_max', 'ISI_mean', 'ISI_median', 'ISI_cv', 'ISI_std',
                       'sparsity', 'selectivity', 'information', 'coherence', 'speed_score', 'hd_score', 'grid_score', 'border_score', 'field_sizes']
    for key in csv_header_keys:
        csv_header[key] = True

    tasks = {}
    task_keys = ['spike_width', 'spike_analysis', 'binary_map', 'autocorrelation_map', 'sparsity', 'selectivity', 'information', 'coherence', 'speed_score', 'hd_score', 'tuning_curve', 'grid_score', 'border_score', 'field_sizes', 'disk_arena']
    for key in task_keys:
        if key == 'disk_arena':
            tasks[key] = False
        else:
            tasks[key] = True

    plotTasks = {}
    plot_task_keys = ['Spikes_Over_Position_Map', 'Tuning_Curve_Plots', 'Firing_Rate_vs_Speed_Plots', 'Firing_Rate_vs_Time_Plots','autocorr_map', 'binary_map','rate_map', 'occupancy_map']
    for key in plot_task_keys:
        if key == 'disk_arena':
            plotTasks[key] = False
        else:
            plotTasks[key] = True

    animal = {'animal_id': '001', 'species': 'mouse', 'sex': 'F', 'age': 1, 'weight': 1, 'genotype': 'type', 'animal_notes': 'notes'}
    devices = {'axona_led_tracker': True, 'implant': True}
    implant = {'implant_id': '001', 'implant_type': 'tetrode', 'implant_geometry': 'square', 'wire_length': 25, 'wire_length_units': 'um', 'implant_units': 'uV'}

    session_settings = {'channel_count': 4, 'animal': animal, 'devices': devices, 'implant': implant}

    settings = {'ppm': 511, 'session':  session_settings, 'smoothing_factor': 3, 'useMatchedCut': True}

    settings['tasks'] = tasks
    settings['plotTasks'] = plotTasks
    settings['header'] = csv_header
    settings['speed_lowerbound'] = 0
    settings['speed_upperbound'] = 99
    settings['end_cell'] = None
    settings['start_cell'] = None
    settings['saveData'] = True
    settings['saveMethod'] = 'one_per_animal'
    # possible saves are:
    # 1 csv per session (all tetrode and indiv): 'one_per_session' --> 5 sheets (summary of all 4 tetrodes, tet1, tet2, tet3, tet4)
    # 1 csv per animal per tetrode (all sessions): 'one_per_animal_tetrode' --> 4 sheets one per tet 
    # 1 csv per animal (all tetrodes & sessions): 'one_per_animal' --> 1 sheet

    start_time = time.time()
    root = tk.Tk()
    root.withdraw()
    data_dir = filedialog.askdirectory(parent=root,title='Please select a data directory.')
    study = make_study(data_dir,settings_dict=settings)
    study.make_animals()

    batch_map(study, settings, data_dir)